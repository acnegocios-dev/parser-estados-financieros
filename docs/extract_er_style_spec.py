"""Extract versioned BG, ER and BAL visual contracts from approved references."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
REFERENCE_DIR = ROOT / "sample-inputs" / "reference-2026-07-20"
SOURCE = REFERENCE_DIR / "EEFF_202602_AL_Serv_Prueba.xlsx"
COLORED_SOURCE = REFERENCE_DIR / "EEFF_202602_AL_Serv_Prueba_referencia_colore_hoja_bal.xlsx"
TARGETS = {
    "ER": ROOT / "src" / "er_style_spec.json",
    "BG": ROOT / "src" / "bg_style_spec.json",
    "BAL": ROOT / "src" / "bal_style_spec.json",
}
BOUNDS = {
    "ER": (1, 70, 1, 10),
    "BG": (1, 47, 1, 12),
    "BAL": (1, 185, 3, 7),
}
ALLOWED_MERGES = {
    "ER": {
        "B9:J9", "B10:J10", "B11:J11", "B12:J12", "B17:C17", "B18:C18",
        "B19:C19", "B20:C20", "B22:C22", "B25:C25", "B27:C27",
        "B51:C51", "B53:C53", "B55:C55", "B58:C58", "B60:C60",
        "B63:C63", "B65:C65", "B67:C67", "B68:C68", "B70:C70",
    },
    "BG": {"B7:L7", "B8:L8", "B9:L9", "B10:L10"},
    "BAL": {"C1:G1", "C2:G2", "C3:G3", "C4:G4"},
}
PAGE_SETUP_FIELDS = (
    "orientation",
    "paperSize",
    "paperHeight",
    "paperWidth",
    "scale",
    "fitToHeight",
    "fitToWidth",
    "firstPageNumber",
    "useFirstPageNumber",
    "pageOrder",
    "blackAndWhite",
    "draft",
    "cellComments",
    "errors",
    "horizontalDpi",
    "verticalDpi",
    "copies",
)
PRINT_OPTION_FIELDS = (
    "gridLines",
    "gridLinesSet",
    "headings",
    "horizontalCentered",
    "verticalCentered",
)
SHEET_VIEW_FIELDS = (
    "showGridLines",
    "topLeftCell",
    "zoomScale",
    "zoomScaleNormal",
    "zoomScalePageLayoutView",
    "view",
)


def _color(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    return {
        "type": value.type,
        "rgb": value.rgb if value.type == "rgb" else None,
        "indexed": value.indexed if value.type == "indexed" else None,
        "theme": value.theme if value.type == "theme" else None,
        "tint": value.tint,
    }


def _side(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    return {"style": value.style, "color": _color(value.color)}


def _style(cell: Any) -> dict[str, Any]:
    font = cell.font
    alignment = cell.alignment
    fill = cell.fill
    border = cell.border
    protection = cell.protection
    return {
        "font": {
            "name": font.name,
            "size": font.sz,
            "bold": font.bold,
            "italic": font.italic,
            "underline": font.underline,
            "strike": font.strike,
            "color": _color(font.color),
            "vertAlign": font.vertAlign,
            "charset": font.charset,
            "family": font.family,
            "scheme": font.scheme,
            "outline": font.outline,
            "shadow": font.shadow,
            "condense": font.condense,
            "extend": font.extend,
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "textRotation": alignment.textRotation,
            "wrapText": alignment.wrap_text,
            "shrinkToFit": alignment.shrink_to_fit,
            "indent": alignment.indent,
            "relativeIndent": alignment.relativeIndent,
            "justifyLastLine": alignment.justifyLastLine,
            "readingOrder": alignment.readingOrder,
        },
        "fill": {
            "fillType": fill.fill_type,
            "fgColor": _color(fill.fgColor),
            "bgColor": _color(fill.bgColor),
        },
        "border": {
            "left": _side(border.left),
            "right": _side(border.right),
            "top": _side(border.top),
            "bottom": _side(border.bottom),
            "start": _side(border.start),
            "end": _side(border.end),
            "diagonal": _side(border.diagonal),
            "diagonalUp": border.diagonalUp,
            "diagonalDown": border.diagonalDown,
            "outline": border.outline,
            "vertical": _side(border.vertical),
            "horizontal": _side(border.horizontal),
        },
        "numberFormat": cell.number_format,
        "protection": {"locked": protection.locked, "hidden": protection.hidden},
    }


def _mask_summary() -> dict[str, Any]:
    workbook = load_workbook(COLORED_SOURCE, data_only=False, keep_links=True)
    worksheet = workbook["BAL"]
    colors: dict[str, list[str]] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.fill.fill_type != "solid" or cell.fill.fgColor.type != "rgb":
                continue
            rgb = cell.fill.fgColor.rgb
            if rgb in {"FFFFFF00", "FF00A933"}:
                colors.setdefault(rgb, []).append(cell.coordinate)
    workbook.close()
    return {
        "source": COLORED_SOURCE.name,
        "source_sha256": hashlib.sha256(COLORED_SOURCE.read_bytes()).hexdigest(),
        "yellow_rgb": "FFFFFF00",
        "yellow_ranges": ["C1:C4", "C5:G185"],
        "yellow_cell_count": len(colors.get("FFFFFF00", [])),
        "green_rgb": "FF00A933",
        "green_ranges": ["H7:M183"],
        "green_cell_count": len(colors.get("FF00A933", [])),
        "output_decision": "Generate only the clean C:G mask; never emit yellow or green annotations.",
    }


def extract(sheet: str) -> dict[str, Any]:
    """Offline extractor only; generated specs are the runtime dependency."""
    workbook = load_workbook(SOURCE, data_only=False, keep_links=True)
    worksheet = workbook[sheet]
    styles: list[dict[str, Any]] = []
    style_ids: dict[str, int] = {}
    cells: dict[str, int] = {}

    first_row, last_row, first_column, last_column = BOUNDS[sheet]
    for row in range(first_row, last_row + 1):
        for column in range(first_column, last_column + 1):
            cell = worksheet.cell(row=row, column=column)
            item = _style(cell)
            key = json.dumps(item, sort_keys=True, ensure_ascii=False)
            if key not in style_ids:
                style_ids[key] = len(styles)
                styles.append(item)
            cells[cell.coordinate] = style_ids[key]

    row_limit = 6 if sheet == "BAL" else last_row
    row_heights = {
        str(row): worksheet.row_dimensions[row].height
        for row in range(1, row_limit + 1)
        if worksheet.row_dimensions[row].height is not None
    }
    columns = [get_column_letter(column) for column in range(first_column, last_column + 1)]
    result = {
        "version": f"2026-07-20.{sheet.lower()}-v2",
        "source": SOURCE.name,
        "source_sha256": hashlib.sha256(SOURCE.read_bytes()).hexdigest(),
        "sheet": sheet,
        "visible_range": (
            f"{worksheet.cell(first_row, first_column).coordinate}:"
            f"{worksheet.cell(last_row, last_column).coordinate}"
        ),
        "geometry": {
            "default_row_height": worksheet.sheet_format.defaultRowHeight,
            "default_column_width": worksheet.sheet_format.defaultColWidth,
            "base_column_width": worksheet.sheet_format.baseColWidth,
            "column_widths": {
                column: worksheet.column_dimensions[column].width for column in columns
            },
            "hidden_columns": [
                column for column in columns if worksheet.column_dimensions[column].hidden
            ],
            "row_heights": row_heights,
            "hidden_rows": [
                row for row in range(1, row_limit + 1) if worksheet.row_dimensions[row].hidden
            ],
            "merged_ranges": sorted(ALLOWED_MERGES[sheet]),
            "page_margins": {
                key: getattr(worksheet.page_margins, key)
                for key in ("left", "right", "top", "bottom", "header", "footer")
            },
            "page_setup": {
                key: getattr(worksheet.page_setup, key)
                for key in PAGE_SETUP_FIELDS
            },
            "page_setup_properties": {
                "fitToPage": worksheet.sheet_properties.pageSetUpPr.fitToPage,
                "autoPageBreaks": worksheet.sheet_properties.pageSetUpPr.autoPageBreaks,
            },
            "print_options": {
                key: getattr(worksheet.print_options, key)
                for key in PRINT_OPTION_FIELDS
            },
            "print_area": str(worksheet.print_area) if worksheet.print_area else None,
            "print_title_rows": worksheet.print_title_rows,
            "print_title_cols": worksheet.print_title_cols,
            "sheet_view": {
                key: getattr(worksheet.sheet_view, key)
                for key in SHEET_VIEW_FIELDS
            },
            "freeze_panes": None,
        },
        "styles": styles,
        "cells": cells,
    }
    if sheet == "BAL":
        result["mask"] = _mask_summary()
        result["dynamic_profiles"] = {
            "light_regular": {"label": cells["C11"], "numeric": cells["D11"]},
            "dark_regular": {"label": cells["C8"], "numeric": cells["D8"]},
            "light_bold": {"label": cells["C7"], "numeric": cells["D7"]},
            "dark_bold": {"label": cells["C23"], "numeric": cells["D23"]},
            "separator": {"label": cells["C184"], "numeric": cells["D184"]},
            "total": {"label": cells["C185"], "numeric": cells["D185"]},
        }
    workbook.close()
    return result


if __name__ == "__main__":
    for sheet, target in TARGETS.items():
        target.write_text(
            json.dumps(extract(sheet), indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        print(target)
