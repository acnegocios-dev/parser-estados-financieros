from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def load_any_xlsx(path: Path):
    with path.open("rb") as fh:
        return load_workbook(fh, data_only=False)


def preview(path: Path) -> None:
    print(f"FILE {path}")
    wb = load_any_xlsx(path)
    print(f"SHEETS {wb.sheetnames}")
    for ws in wb.worksheets:
        print(f"SHEET {ws.title} {ws.max_row}x{ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            print(row)
        print("")


def preview_non_empty(path: Path, sheet_name: str, max_rows: int = 120) -> None:
    print(f"NON_EMPTY {path.name} {sheet_name}")
    wb = load_any_xlsx(path)
    ws = wb[sheet_name]
    shown = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = row[:11]
        if any(value is not None for value in cells):
            print(idx, cells)
            shown += 1
        if shown >= max_rows:
            break


def preview_rows(path: Path, sheet_name: str, first_row: int, last_row: int) -> None:
    print(f"ROWS {path.name} {sheet_name} {first_row}:{last_row}")
    wb = load_any_xlsx(path)
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=first_row, max_row=last_row, values_only=True):
        print(row)


if __name__ == "__main__":
    preview(ROOT / "sample-inputs" / "balanza_SME170717GA0_2026_07.xls")
    preview(ROOT / "sample-inputs" / "EEFF_202602_AL_Serv_Prueba.xlsx")
    preview_non_empty(ROOT / "sample-inputs" / "EEFF_202602_AL_Serv_Prueba.xlsx", "ER")
    preview_rows(ROOT / "sample-inputs" / "balanza_SME170717GA0_2026_07.xls", "Balanza", 90, 150)
    preview_rows(
        ROOT / "sample-outputs" / "estado_resultados_al_servicios_multiples_empresariales_sa_de_cv_2026_07.xlsx",
        "ER",
        17,
        70,
    )
