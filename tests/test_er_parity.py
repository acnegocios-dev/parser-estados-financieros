from __future__ import annotations

import hashlib
import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.engine import build_er_dataset
from src.parser import parse_balanza
from src.validation import validate_generated_workbook
from src.workbook import ER_LAYOUT, build_er_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "sample-inputs" / "balanza_SME170717GA0_2026_07.xls"
# This historical workbook is used only for cached-value regression.  Its
# styles are intentionally never compared or copied into the generated ER.
CACHED_VALUE_REGRESSION_MANUAL = ROOT / "sample-inputs" / "EEFF_202602_AL_Serv_Prueba.xlsx"
# This approved SHA-256 reference is the sole visual/geometry source.
VISUAL_STYLE_REFERENCE = (
    ROOT
    / "sample-inputs"
    / "reference-2026-07-20"
    / "EEFF_202602_AL_Serv_Prueba.xlsx"
)
STYLE_SPEC = ROOT / "src" / "er_style_spec.json"
VISUAL_STYLE_SHA256 = "991daeaa5b9f957e490e231164825640127cb850f01db865c04cfbb25e72b12c"
CACHED_VALUE_SHA256 = "d530b541450f6fa9b1bfbece5c8cf4d811b97096884587edfde92156ef4ce0cc"
SUBTOTAL_ROWS = {20, 25, 51, 53, 58, 63, 65, 70}
KEY_STYLE_CELLS = (
    "B9", "B10", "B11", "B12", "D15", "F15", "H15", "J15",
    "B17", "B18", "H18", "J18", "B20", "H20", "B22", "B23",
    "H23", "B25", "H25", "B27", "B28", "B46", "H46", "B51",
    "H51", "B55", "B56", "B60", "B61", "H61", "B62", "H62",
    "B63", "H63", "B65", "H65", "B67", "H67", "B68", "H68",
    "B70", "H70", "J70",
)


def _build_workbook():
    parsed = parse_balanza(INPUT)
    dataset = build_er_dataset(
        [row.to_dict() for row in parsed.rows],
        company=parsed.company_name,
        period=parsed.period.period_ym,
        source_path=parsed.source_path,
    )
    return build_er_workbook(dataset, source_path=parsed.source_path).workbook


def _color_signature(color):
    if color is None:
        return None
    return (
        color.type,
        color.rgb if color.type == "rgb" else None,
        color.indexed if color.type == "indexed" else None,
        color.theme if color.type == "theme" else None,
        color.tint,
    )


def _side_signature(side):
    if side is None:
        return None
    return side.style, _color_signature(side.color)


def _style_signature(cell):
    font = cell.font
    alignment = cell.alignment
    fill = cell.fill
    border = cell.border
    return (
        (font.name, font.sz, font.bold, font.italic, font.underline, font.strike, _color_signature(font.color)),
        (
            alignment.horizontal,
            alignment.vertical,
            alignment.textRotation,
            alignment.wrap_text,
            alignment.shrink_to_fit,
            alignment.indent,
            alignment.relativeIndent,
            alignment.justifyLastLine,
            alignment.readingOrder,
        ),
        (
            fill.fill_type,
            _color_signature(fill.fgColor),
            _color_signature(fill.bgColor),
        ),
        (
            _side_signature(border.left),
            _side_signature(border.right),
            _side_signature(border.top),
            _side_signature(border.bottom),
            _side_signature(border.diagonal),
            border.diagonalUp,
            border.diagonalDown,
            border.outline,
            _side_signature(border.vertical),
            _side_signature(border.horizontal),
        ),
        cell.number_format,
        (cell.protection.locked, cell.protection.hidden),
    )


class ErWorkbookParityTest(unittest.TestCase):
    def test_cached_value_and_visual_style_references_are_explicitly_separated(self) -> None:
        self.assertEqual(
            hashlib.sha256(VISUAL_STYLE_REFERENCE.read_bytes()).hexdigest(), VISUAL_STYLE_SHA256
        )
        self.assertEqual(
            hashlib.sha256(CACHED_VALUE_REGRESSION_MANUAL.read_bytes()).hexdigest(), CACHED_VALUE_SHA256
        )
        visual_values = load_workbook(VISUAL_STYLE_REFERENCE, data_only=True, keep_links=True)["ER"]
        cached_values = load_workbook(
            CACHED_VALUE_REGRESSION_MANUAL, data_only=True, keep_links=True
        )["ER"]
        for coordinate in ("H18", "H46", "H70"):
            self.assertIsNone(visual_values[coordinate].value, coordinate)
            self.assertIsNotNone(cached_values[coordinate].value, coordinate)

    def test_detail_amounts_match_manual_cached_values_with_cent_tolerance(self) -> None:
        generated = _build_workbook()["ER"]
        manual = load_workbook(
            CACHED_VALUE_REGRESSION_MANUAL, data_only=True, keep_links=True
        )["ER"]

        detail_rows = [
            int(spec["row"])
            for spec in ER_LAYOUT
            if spec.get("kind") not in {"section", "subtotal"}
        ]
        differences = []
        for row in detail_rows:
            actual = float(generated[f"H{row}"].value or 0)
            expected = float(manual[f"H{row}"].value or 0)
            difference = actual - expected
            if abs(difference) > 0.01:
                differences.append((row, actual, expected, difference))
        self.assertEqual(differences, [])

    def test_h46_uses_the_three_manual_composite_prefixes(self) -> None:
        from src.engine import ER_MAPPED_LINES

        varios = next(line for line in ER_MAPPED_LINES if line["key"] == "varios")
        self.assertEqual(set(varios["codes"]), {"6148", "6176", "6195"})
        self.assertAlmostEqual(_build_workbook()["ER"]["H46"].value, 39614.91)

    def test_sheet_geometry_matches_manual(self) -> None:
        generated = _build_workbook()["ER"]
        manual = load_workbook(VISUAL_STYLE_REFERENCE, data_only=False, keep_links=True)["ER"]

        self.assertIsNone(generated.freeze_panes)
        self.assertEqual(generated.sheet_format.defaultRowHeight, manual.sheet_format.defaultRowHeight)
        self.assertEqual(generated.sheet_format.defaultColWidth, manual.sheet_format.defaultColWidth)
        self.assertEqual(generated.sheet_format.baseColWidth, manual.sheet_format.baseColWidth)
        self.assertEqual(
            {column: generated.column_dimensions[column].width for column in "ABCDEFGHIJ"},
            {column: manual.column_dimensions[column].width for column in "ABCDEFGHIJ"},
        )
        self.assertEqual(
            {row: generated.row_dimensions[row].height for row in range(1, 71)},
            {row: manual.row_dimensions[row].height for row in range(1, 71)},
        )
        self.assertEqual(
            [row for row in range(1, 71) if generated.row_dimensions[row].hidden],
            [row for row in range(1, 71) if manual.row_dimensions[row].hidden],
        )
        manual_merges = {
            str(value)
            for value in manual.merged_cells.ranges
            if value.max_row <= 70 and value.max_col <= 10
        }
        self.assertEqual(
            {str(value) for value in generated.merged_cells.ranges},
            manual_merges,
        )
        self.assertEqual(
            {column: generated.column_dimensions[column].hidden for column in "ABCDEFGHIJ"},
            {column: manual.column_dimensions[column].hidden for column in "ABCDEFGHIJ"},
        )
        self.assertEqual(generated.page_setup.orientation, manual.page_setup.orientation)
        self.assertEqual(generated.page_setup.paperSize, manual.page_setup.paperSize)
        self.assertEqual(generated.page_setup.horizontalDpi, manual.page_setup.horizontalDpi)
        self.assertEqual(generated.page_setup.verticalDpi, manual.page_setup.verticalDpi)
        for name in ("left", "right", "top", "bottom", "header", "footer"):
            self.assertEqual(getattr(generated.page_margins, name), getattr(manual.page_margins, name))
        self.assertEqual(generated.print_area, "'ER'!$B$9:$J$70")
        self.assertEqual(generated.print_title_rows, "$9:$15")
        self.assertEqual(generated.page_setup.fitToWidth, 1)
        self.assertEqual(generated.page_setup.fitToHeight, 0)

    def test_key_non_fill_styles_match_manual_and_effective_area_is_white(self) -> None:
        generated = _build_workbook()["ER"]
        manual = load_workbook(VISUAL_STYLE_REFERENCE, data_only=False, keep_links=True)["ER"]
        for coordinate in KEY_STYLE_CELLS:
            with self.subTest(coordinate=coordinate):
                generated_style = _style_signature(generated[coordinate])
                manual_style = _style_signature(manual[coordinate])
                self.assertEqual(
                    generated_style[:2] + generated_style[3:],
                    manual_style[:2] + manual_style[3:],
                )
                self.assertEqual(generated[coordinate].fill.fill_type, "solid")
                self.assertEqual(generated[coordinate].fill.fgColor.rgb, "FFFFFFFF")

    def test_style_spec_is_versioned_and_self_contained(self) -> None:
        spec = json.loads(STYLE_SPEC.read_text(encoding="utf-8"))
        self.assertEqual(spec["version"], "2026-07-20.er-v2")
        self.assertEqual(spec["source"], VISUAL_STYLE_REFERENCE.name)
        self.assertEqual(spec["source_sha256"], VISUAL_STYLE_SHA256)
        self.assertIn("B9", spec["cells"])
        fonts = {style["font"]["name"] for style in spec["styles"]}
        self.assertEqual(fonts, {"Arial"})

    def test_workbook_has_no_external_links_or_formula_error_tokens(self) -> None:
        workbook = _build_workbook()
        validation = validate_generated_workbook(workbook, balance_difference=0.0)
        self.assertTrue(validation.formula_static_validation)
        self.assertIsNone(validation.formula_evaluated_error_count)
        self.assertEqual(getattr(workbook, "_external_links", []), [])
        tokens = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        self.assertFalse(any(token in cell.value.upper() for token in tokens), cell.coordinate)

    def test_percentages_use_safe_if_formulas_and_period_columns_stay_without_formulas(self) -> None:
        worksheet = _build_workbook()["ER"]
        numeric_rows = [
            int(spec["row"])
            for spec in ER_LAYOUT
            if spec.get("kind") != "section"
        ]
        for row in numeric_rows:
            self.assertEqual(worksheet[f"J{row}"].value, f"=IF($H$18=0,0,H{row}/$H$18)")
        for row in range(1, 71):
            for column in ("D", "F"):
                value = worksheet[f"{column}{row}"].value
                self.assertFalse(isinstance(value, str) and value.startswith("="), f"{column}{row}")


if __name__ == "__main__":
    unittest.main()
