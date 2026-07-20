from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import load_workbook

from src.engine import build_bg_dataset, build_er_dataset
from src.parser import parse_balanza
from src.validation import validate_balance_sheet
from src.workbook import build_er_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "sample-inputs" / "balanza_SME170717GA0_2026_07.xls"
MANUAL = (
    ROOT
    / "sample-inputs"
    / "reference-2026-07-20"
    / "EEFF_202602_AL_Serv_Prueba.xlsx"
)
BG_STYLE_SPEC = ROOT / "src" / "bg_style_spec.json"
BAL_STYLE_SPEC = ROOT / "src" / "bal_style_spec.json"


def _build():
    parsed = parse_balanza(INPUT)
    dataset = build_er_dataset(
        [row.to_dict() for row in parsed.rows],
        company=parsed.company_name,
        period=parsed.period.period_ym,
        source_path=parsed.source_path,
    )
    return parsed, build_er_workbook(dataset, source_path=parsed.source_path).workbook


def _color(color):
    if color is None:
        return None
    return (
        color.type,
        color.rgb if color.type == "rgb" else None,
        color.indexed if color.type == "indexed" else None,
        color.theme if color.type == "theme" else None,
        color.tint,
    )


def _side(side):
    if side is None:
        return None
    return side.style, _color(side.color)


def _style(cell, include_fill=True):
    font = cell.font
    alignment = cell.alignment
    border = cell.border
    result = (
        (
            font.name,
            font.sz,
            font.bold,
            font.italic,
            font.underline,
            font.strike,
            _color(font.color),
        ),
        (
            alignment.horizontal,
            alignment.vertical,
            alignment.textRotation,
            alignment.wrap_text,
            alignment.shrink_to_fit,
            alignment.indent,
        ),
        (
            _side(border.left),
            _side(border.right),
            _side(border.top),
            _side(border.bottom),
        ),
        cell.number_format,
        (cell.protection.locked, cell.protection.hidden),
    )
    if not include_fill:
        return result
    fill = (
        cell.fill.fill_type,
        _color(cell.fill.fgColor),
        _color(cell.fill.bgColor),
    )
    return result[:2] + (fill,) + result[2:]


class BgBalParityTest(unittest.TestCase):
    def test_bal_uses_only_versioned_profiles_and_clean_runtime_contract(self):
        parsed, workbook = _build()
        bal = workbook["BAL"]
        spec = json.loads(BAL_STYLE_SPEC.read_text(encoding="utf-8"))

        self.assertEqual(spec["visible_range"], "C1:G185")
        self.assertEqual(spec["mask"]["source_sha256"], "c27d3b4f40737e00a01dc83a2bc8745f6d62fd4a17cd8d78600a5ec09764dda2")
        self.assertEqual(set(spec["dynamic_profiles"]), {
            "light_regular", "dark_regular", "light_bold", "dark_bold", "separator", "total",
        })
        self.assertEqual(len(parsed.rows), 157)
        self.assertEqual(bal.max_row, 165)
        self.assertIsNone(bal.freeze_panes)
        self.assertEqual(bal.print_area, "'BAL'!$C$1:$G$165")
        self.assertEqual(bal.print_title_rows, "$1:$6")
        self.assertEqual(bal.page_setup.orientation, "portrait")
        self.assertEqual(bal.page_setup.paperSize, 9)
        self.assertEqual(bal.page_setup.fitToWidth, 1)
        self.assertEqual(bal.page_setup.fitToHeight, 0)
        self.assertTrue(bal.sheet_properties.pageSetUpPr.fitToPage)
        for row in range(1, 166):
            for column in (1, 2, 8, 9, 10, 11, 12, 13):
                self.assertIsNone(bal.cell(row, column).value)

    def test_bg_uses_the_versioned_spec_and_safe_print_contract(self):
        _, workbook = _build()
        generated = workbook["BG"]
        spec = json.loads(BG_STYLE_SPEC.read_text(encoding="utf-8"))

        self.assertEqual(spec["visible_range"], "A1:L47")
        self.assertEqual(spec["source_sha256"], "991daeaa5b9f957e490e231164825640127cb850f01db865c04cfbb25e72b12c")
        self.assertEqual(
            {str(value) for value in generated.merged_cells.ranges},
            set(spec["geometry"]["merged_ranges"]),
        )
        self.assertEqual(generated.column_dimensions["A"].width, spec["geometry"]["column_widths"]["A"])
        self.assertEqual(generated.column_dimensions["L"].width, spec["geometry"]["column_widths"]["L"])
        self.assertEqual(generated["B7"].font.name, spec["styles"][spec["cells"]["B7"]]["font"]["name"])
        self.assertEqual(generated["L47"].number_format, spec["styles"][spec["cells"]["L47"]]["numberFormat"])
        self.assertEqual(generated.print_area, "'BG'!$B$7:$L$47")
        self.assertEqual(generated.print_title_rows, "$7:$10")
        self.assertEqual(generated.page_setup.orientation, "portrait")
        self.assertEqual(generated.page_setup.paperSize, 9)
        self.assertEqual(generated.page_setup.fitToWidth, 1)
        self.assertEqual(generated.page_setup.fitToHeight, 0)
        self.assertTrue(generated.sheet_properties.pageSetUpPr.fitToPage)

        expected_formulas = {
            "F26": "=SUM(E16:E24)",
            "F36": "=SUM(E29:E34)",
            "F42": "=SUM(E39:E40)",
            "F45": "=F26+F36+F42",
            "L26": "=SUM(K16:K20)",
            "K34": "=ER!H70",
            "L36": "=SUM(K31:K34)",
            "L45": "=L26+L36",
            "L47": "=F45-L45",
        }
        self.assertEqual({cell: generated[cell].value for cell in expected_formulas}, expected_formulas)
        for formula in expected_formulas.values():
            self.assertNotIn("BAL!", formula)
            self.assertNotIn("/4", formula)
            self.assertNotIn("-0.03", formula)
            self.assertNotIn("==", formula)

    def test_bg_l47_programmatic_evidence_uses_the_same_account_dataset(self):
        parsed, workbook = _build()
        er = build_er_dataset(parsed.rows)
        bg = build_bg_dataset(
            parsed.rows,
            result_ejercicio=er["raw_amounts"]["resultado_ejercicio"],
        )
        validation = validate_balance_sheet(
            parsed.rows,
            result_ejercicio=er["raw_amounts"]["resultado_ejercicio"],
        )

        self.assertEqual(workbook["BG"]["L47"].value, "=F45-L45")
        self.assertEqual(bg["formula"], "F45-L45")
        self.assertAlmostEqual(bg["diferencia_cuadre"], validation.diferencia_cuadre, places=8)
        self.assertAlmostEqual(bg["diferencia_cuadre"], -0.059663, places=6)
        self.assertLess(abs(bg["diferencia_cuadre"]), 1)

    def test_bg_geometry_and_non_fill_styles_match_effective_manual_area(self):
        _, workbook = _build()
        generated = workbook["BG"]
        manual = load_workbook(MANUAL, data_only=False, keep_links=True)["BG"]

        self.assertEqual(
            {column: generated.column_dimensions[column].width for column in "ABCDEFGHIJKL"},
            {column: manual.column_dimensions[column].width for column in "ABCDEFGHIJKL"},
        )
        self.assertEqual(
            {row: generated.row_dimensions[row].height for row in range(1, 48)},
            {row: manual.row_dimensions[row].height for row in range(1, 48)},
        )
        self.assertEqual(
            [row for row in range(1, 48) if generated.row_dimensions[row].hidden],
            [row for row in range(1, 48) if manual.row_dimensions[row].hidden],
        )
        self.assertEqual(
            {str(value) for value in generated.merged_cells.ranges},
            {"B7:L7", "B8:L8", "B9:L9", "B10:L10"},
        )
        key_cells = (
            "B7", "B8", "B9", "B10", "C13", "I13", "C14", "I14",
            "B16", "E16", "H16", "K16", "B26", "F26", "C28",
            "I29", "B31", "E31", "H31", "K31", "B36", "F36",
            "H36", "L36", "B45", "F45", "H45", "L45", "L47",
        )
        for coordinate in key_cells:
            with self.subTest(coordinate=coordinate):
                self.assertEqual(
                    _style(generated[coordinate], include_fill=False),
                    _style(manual[coordinate], include_fill=False),
                )
        for row in generated.iter_rows(min_row=1, max_row=47, min_col=1, max_col=12):
            for cell in row:
                self.assertEqual(cell.fill.fill_type, "solid")
                self.assertEqual(cell.fill.fgColor.rgb, "FFFFFFFF")

    def test_bal_uses_clean_mask_styles_as_dynamic_profiles(self):
        parsed, workbook = _build()
        generated = workbook["BAL"]
        manual = load_workbook(MANUAL, data_only=False, keep_links=True)["BAL"]

        self.assertEqual(
            {column: generated.column_dimensions[column].width for column in "CDEFG"},
            {column: manual.column_dimensions[column].width for column in "CDEFG"},
        )
        for coordinate in ("C1", "C2", "C3", "C4", "C5", "C6", "D6", "E6", "F6", "G6"):
            with self.subTest(coordinate=coordinate):
                self.assertEqual(_style(generated[coordinate]), _style(manual[coordinate]))

        references = {
            ("light", False): ("C11", "D11"),
            ("dark", False): ("C8", "D8"),
            ("light", True): ("C7", "D7"),
            ("dark", True): ("C23", "D23"),
        }
        for row_number, source_row in enumerate(parsed.rows, start=7):
            band = "light" if (row_number - 7) % 2 == 0 else "dark"
            label_ref, numeric_ref = references[(band, "-" not in source_row.account_code)]
            self.assertEqual(_style(generated.cell(row_number, 3)), _style(manual[label_ref]))
            for column in range(4, 8):
                self.assertEqual(_style(generated.cell(row_number, column)), _style(manual[numeric_ref]))

        for generated_row, reference_row in ((164, 184), (165, 185)):
            self.assertEqual(
                _style(generated.cell(generated_row, 3)),
                _style(manual.cell(reference_row, 3)),
            )
            for column in range(4, 8):
                self.assertEqual(
                    _style(generated.cell(generated_row, column)),
                    _style(manual.cell(reference_row, 4)),
                )

    def test_approved_mask_colors_never_reach_the_output(self):
        _, workbook = _build()
        forbidden = {"FFFFFF00", "FF00A933"}
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    color = cell.fill.fgColor
                    if cell.fill.fill_type and color.type == "rgb":
                        self.assertNotIn(color.rgb, forbidden, cell.coordinate)


if __name__ == "__main__":
    unittest.main()
