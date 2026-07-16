from __future__ import annotations

import unittest
from pathlib import Path
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from src.engine import build_bal_dataset, build_bg_dataset, build_er_dataset
from src.parser import parse_balanza
from src.validation import validate_financial_statements_workbook
from src.workbook import build_financial_statements_workbook


ROOT = Path(__file__).resolve().parents[1]
SAMPLE = ROOT / "sample-inputs" / "balanza_SME170717GA0_2026_07.xls"


class FinancialStatementsWorkbookValidationTest(unittest.TestCase):
    """Runtime-contract checks: deliberately no manual workbook is opened here."""

    def _build(self):
        parsed = parse_balanza(SAMPLE)
        er = build_er_dataset(
            parsed.rows,
            company=parsed.company_name,
            period=parsed.period.period_ym,
            source_path=parsed.source_path,
        )
        bg = build_bg_dataset(parsed.rows, result_ejercicio=er["raw_amounts"]["resultado_ejercicio"])
        bal = build_bal_dataset(parsed.rows)
        result = build_financial_statements_workbook(
            er,
            bg_dataset=bg,
            bal_dataset=bal,
            metadata={
                "company": parsed.company_name,
                "period": parsed.period.period_ym,
                "rfc": parsed.period.rfc,
            },
            source_path=parsed.source_path,
        )
        return parsed, bg, bal, result.workbook

    def test_static_contract_and_programmatic_evidence_close(self) -> None:
        parsed, bg, bal, workbook = self._build()

        result = validate_financial_statements_workbook(
            workbook,
            expected_company=parsed.company_name,
            expected_period=parsed.period.period_ym,
            expected_rfc=parsed.period.rfc,
            normalized_rows=len(parsed.rows),
            bal_dataset=bal,
            bg_dataset=bg,
        )

        self.assertTrue(result.ok, result.issues)
        self.assertTrue(result.formula_validation.formula_static_validation)
        self.assertFalse(result.formula_validation.formula_recalculation_performed)
        self.assertIsNone(result.formula_validation.formula_evaluated_error_count)
        self.assertFalse(result.formula_validation.formula_cached_values_available)
        self.assertEqual(result.evidence["bal_sumas_iguales"], {
            "saldo_inicial": 0.0,
            "debe": 584.64,
            "haber": 584.64,
            "saldo_final": 0.0,
        })
        self.assertAlmostEqual(result.evidence["bg_balance"]["difference"], -0.059663, places=6)
        self.assertLess(abs(result.evidence["bg_balance"]["difference"]), 1)
        self.assertEqual(workbook["ER"]["H46"].value, 39614.91)

    def test_rejects_a_mask_fill_and_bal_data_outside_table(self) -> None:
        parsed, bg, bal, workbook = self._build()
        workbook["BAL"]["H7"] = "residuo"
        workbook["BAL"]["C7"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")

        result = validate_financial_statements_workbook(
            workbook,
            expected_company=parsed.company_name,
            expected_period=parsed.period.period_ym,
            expected_rfc=parsed.period.rfc,
            normalized_rows=len(parsed.rows),
            bal_dataset=bal,
            bg_dataset=bg,
        )

        self.assertFalse(result.ok)
        self.assertTrue(any("outside C:G" in issue for issue in result.issues))
        self.assertTrue(any("forbidden mask fill" in issue for issue in result.issues))

    def test_serialized_merged_titles_keep_an_effective_white_fill(self) -> None:
        parsed, bg, bal, workbook = self._build()
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "estados_financieros.xlsx"
            workbook.save(output)
            reloaded = load_workbook(output)
            result = validate_financial_statements_workbook(
                reloaded,
                expected_company=parsed.company_name,
                expected_period=parsed.period.period_ym,
                expected_rfc=parsed.period.rfc,
                normalized_rows=len(parsed.rows),
                bal_dataset=bal,
                bg_dataset=bg,
            )

        self.assertTrue(result.ok, result.issues)


if __name__ == "__main__":
    unittest.main()
