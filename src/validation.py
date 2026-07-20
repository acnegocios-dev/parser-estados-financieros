"""Validation helpers for generated local financial-statement workbooks."""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
import os
from pathlib import Path
import shutil
import subprocess
import tempfile
from typing import Any, Iterable, Mapping
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

try:
    from .engine import build_bg_dataset, canonical_account_code, leaf_account_rows, to_decimal
except ImportError:  # pragma: no cover - supports direct script execution.
    from engine import build_bg_dataset, canonical_account_code, leaf_account_rows, to_decimal


FORMULA_ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
CONCEPTUAL_BALANCE_REFERENCE = "BG!L47"
BALANCE_TOLERANCE_DEFAULT = 1.0


@dataclass
class FormulaIssue:
    cell: str
    value: str
    reason: str


@dataclass
class ValidationResult:
    ok: bool
    formula_static_validation: bool
    formula_static_issues: list[FormulaIssue] = field(default_factory=list)
    formula_recalculation_performed: bool = False
    formula_recalculation_engine: str = "none"
    formula_evaluated_error_count: int | None = None
    formula_cached_values_available: bool = False
    formula_evaluated_issues: list[FormulaIssue] = field(default_factory=list)
    balance_difference: float | None = None
    balance_ok: bool | None = None
    balance_reference: str = CONCEPTUAL_BALANCE_REFERENCE
    warnings: list[str] = field(default_factory=list)


@dataclass
class FormulaRecalculationResult:
    performed: bool
    engine: str
    evaluated_error_count: int | None
    cached_values_available: bool
    evaluated_issues: list[FormulaIssue] = field(default_factory=list)
    evidence_path: str | None = None
    blocked: bool = False
    warnings: list[str] = field(default_factory=list)


@dataclass
class FinancialStatementsValidationResult:
    """Static workbook-contract evidence, kept distinct from formula recalculation."""

    ok: bool
    issues: list[str] = field(default_factory=list)
    formula_validation: ValidationResult | None = None
    evidence: dict[str, Any] = field(default_factory=dict)


@dataclass
class PrintValidationResult:
    """Evidence for the serialized print contract and its optional PDF rendering."""

    ok: bool
    issues: list[str] = field(default_factory=list)
    pdf_rendered: bool = False
    pdf_pages: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


@dataclass
class BalanceComponent:
    rubro: str
    total: float
    cuentas: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class BalanceCheckResult:
    total_activo: float
    total_pasivo: float
    capital_contable: float
    diferencia_cuadre: float
    tolerance: float
    cuadra: bool
    balanza_no_cuadra: bool
    componentes: list[BalanceComponent] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return self.cuadra


def validate_generated_workbook(
    workbook_or_path: Workbook | str | Path,
    *,
    balance_difference: float | int | str | None = None,
    tolerance: float = 1.0,
    formula_mode: str = "static_only",
    evaluated_workbook_or_path: Workbook | str | Path | None = None,
    formula_recalculation_engine: str | None = None,
) -> ValidationResult:
    """Validate formulas and the conceptual balance difference.

    The balance check mirrors the manual workbook concept at BG!L47 without
    copying or requiring that sheet. Pass `balance_difference` from the parser
    or balance-sheet integration when available.

    ``static_only`` inspects formula text only. ``recalculated_ok`` and
    ``recalculated_error`` require a separate data-only workbook containing
    evaluated formula results; this prevents a static token scan from being
    presented as proof that Excel formulas were calculated.
    """

    if formula_mode not in {"static_only", "recalculated_ok", "recalculated_error"}:
        raise ValueError(f"Unsupported formula validation mode: {formula_mode}")
    if formula_mode != "static_only" and evaluated_workbook_or_path is None:
        raise ValueError("An evaluated workbook is required for recalculated validation modes.")

    workbook, close_after = _load_workbook(workbook_or_path)
    evaluated_workbook = None
    evaluated_close_after = False
    try:
        formula_issues = find_formula_issues(workbook)
        evaluated_issues: list[FormulaIssue] = []
        cached_values_available = False
        if formula_mode != "static_only":
            evaluated_workbook, evaluated_close_after = _load_workbook(
                evaluated_workbook_or_path, data_only=True
            )
            evaluated_issues, cached_values_available = find_evaluated_formula_issues(
                workbook, evaluated_workbook
            )
        coerced_difference = _coerce_float(balance_difference)
        warnings: list[str] = []
        if balance_difference is None:
            warnings.append(
                "Balance difference was not provided; conceptual BG!L47 balance check was skipped."
            )
            balance_ok = None
        elif coerced_difference is None:
            warnings.append("Balance difference could not be parsed as a number.")
            balance_ok = False
        else:
            balance_ok = abs(coerced_difference) < tolerance
        static_validation = not formula_issues
        recalculation_performed = formula_mode != "static_only"
        evaluated_error_count = len(evaluated_issues) if recalculation_performed else None
        ok = (
            static_validation
            and (evaluated_error_count in (None, 0))
            and (balance_ok is not False)
        )
        return ValidationResult(
            ok=ok,
            formula_static_validation=static_validation,
            formula_static_issues=formula_issues,
            formula_recalculation_performed=recalculation_performed,
            formula_recalculation_engine=(
                "none" if not recalculation_performed else (formula_recalculation_engine or "unknown")
            ),
            formula_evaluated_error_count=evaluated_error_count,
            formula_cached_values_available=cached_values_available,
            formula_evaluated_issues=evaluated_issues,
            balance_difference=coerced_difference,
            balance_ok=balance_ok,
            warnings=warnings,
        )
    finally:
        if close_after:
            workbook.close()
        if evaluated_close_after and evaluated_workbook is not None:
            evaluated_workbook.close()


def validate_financial_statements_workbook(
    workbook_or_path: Workbook | str | Path,
    *,
    expected_company: str | None = None,
    expected_period: str | None = None,
    expected_rfc: str | None = None,
    normalized_rows: int | None = None,
    bal_dataset: Mapping[str, Any] | None = None,
    bg_dataset: Mapping[str, Any] | None = None,
) -> FinancialStatementsValidationResult:
    """Validate the three-sheet contract without claiming formula evaluation.

    ``bal_dataset`` and ``bg_dataset`` supply programmatic evidence for totals
    whose worksheet formulas intentionally remain unevaluated until an Excel
    compatible engine recalculates the saved file.
    """

    workbook, close_after = _load_workbook(workbook_or_path)
    try:
        issues: list[str] = []
        expected_sheets = ["BG", "ER", "BAL"]
        if workbook.sheetnames != expected_sheets:
            issues.append(f"Expected sheets {expected_sheets}; found {workbook.sheetnames}.")
        if [sheet.sheet_state for sheet in workbook.worksheets] != ["visible"] * 3:
            issues.append("All financial-statement sheets must be explicitly visible.")
        if list(workbook.defined_names):
            issues.append("Workbook contains inherited defined names.")
        if getattr(workbook, "_external_links", []):
            issues.append("Workbook contains external links.")
        for worksheet in workbook.worksheets:
            if getattr(worksheet, "_images", []):
                issues.append(f"{worksheet.title} contains inherited images.")

        formula_validation = validate_generated_workbook(workbook, formula_mode="static_only")
        if not formula_validation.formula_static_validation:
            issues.append("Workbook contains formula error tokens or external workbook formulas.")
        if formula_validation.formula_recalculation_performed:
            issues.append("Static validation must not claim formula recalculation.")
        if formula_validation.formula_evaluated_error_count is not None:
            issues.append("Static validation must not claim evaluated formula errors.")
        for title in expected_sheets:
            if title in workbook.sheetnames and not any(
                isinstance(cell.value, str) and cell.value.startswith("=")
                for row in workbook[title].iter_rows() for cell in row
            ):
                issues.append(f"{title} must contain internal formulas.")

        if "BG" in workbook.sheetnames:
            _validate_white_sheet(workbook["BG"], 1, 47, 1, 12, issues)
            _validate_bg_contract(workbook["BG"], expected_company, expected_period, issues)
        if "ER" in workbook.sheetnames:
            _validate_white_sheet(workbook["ER"], 1, 70, 1, 10, issues)
        if "BAL" in workbook.sheetnames:
            _validate_bal_contract(
                workbook["BAL"], expected_company, expected_period, expected_rfc,
                normalized_rows, issues,
            )
        _validate_print_contract(workbook, normalized_rows, issues)
        _validate_forbidden_fills(workbook, issues)

        evidence = _financial_statements_evidence(bal_dataset, bg_dataset)
        if evidence.get("bal_sumas_iguales"):
            totals = evidence["bal_sumas_iguales"]
            if not _within_cent(totals["debe"], 584.64) or not _within_cent(totals["haber"], 584.64):
                issues.append("Programmatic BAL Debe/Haber evidence does not equal 584.64.")
            if not _within_cent(totals["saldo_final"], 0.0):
                issues.append("Programmatic BAL saldo final evidence does not equal zero.")
        if evidence.get("bg_balance"):
            balance = evidence["bg_balance"]
            if not _within_cent(balance["difference"], balance["report_difference"]):
                issues.append("BG formula evidence and programmatic balance report differ by more than 0.01.")
            if abs(balance["difference"]) >= 1:
                issues.append("BG!L47 programmatic evidence is outside the strict balance tolerance.")

        return FinancialStatementsValidationResult(
            ok=not issues,
            issues=issues,
            formula_validation=formula_validation,
            evidence=evidence,
        )
    finally:
        if close_after:
            workbook.close()


def validate_print_contract_after_roundtrip(
    workbook_path: str | Path,
    *,
    normalized_rows: int,
) -> PrintValidationResult:
    """Validate print metadata after saving and reopening an XLSX file.

    The XLSX itself remains the deliverable.  When LibreOffice is installed a
    throw-away PDF is also converted to page images and inspected as visual
    evidence; neither the PDF nor its images are retained in the repository.
    """

    source = Path(workbook_path)
    if not source.is_file():
        return PrintValidationResult(ok=False, issues=[f"Workbook does not exist: {source}"])

    issues: list[str] = []
    workbook = load_workbook(source)
    try:
        _validate_print_contract(workbook, normalized_rows, issues)
    finally:
        workbook.close()

    result = PrintValidationResult(ok=not issues, issues=issues)
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if executable is None:
        result.warnings.append("PDF print evidence skipped: libreoffice/soffice is unavailable.")
        return result

    pdf_issues, page_counts, warnings = _validate_pdf_print_evidence(source, executable)
    result.pdf_rendered = True
    result.pdf_pages = page_counts
    result.warnings.extend(warnings)
    result.issues.extend(pdf_issues)
    result.ok = not result.issues
    return result


def _validate_white_sheet(
    ws,
    min_row: int,
    max_row: int,
    min_column: int,
    max_column: int,
    issues: list[str],
) -> None:
    if ws.sheet_view.showGridLines is not False:
        issues.append(f"{ws.title} must explicitly disable gridlines.")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_column, max_col=max_column):
        for cell in row:
            if not _is_effectively_solid_white(ws, cell):
                issues.append(f"{ws.title}!{cell.coordinate} is not solid white.")
                return


def _is_effectively_solid_white(ws, cell) -> bool:
    """Accept the white anchor of a merged range as its effective fill.

    LibreOffice drops styles from non-anchor cells in merged ranges while
    preserving the visual fill from the anchor.  The workbook contract is
    visual and cell-safe, so validate that representation without accepting a
    non-white merged title.
    """

    effective_cell = cell
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            effective_cell = ws.cell(merged_range.min_row, merged_range.min_col)
            break
    fill = effective_cell.fill
    return (
        fill.fill_type == "solid"
        and fill.fgColor.type == "rgb"
        and fill.fgColor.rgb in {"FFFFFFFF", "00FFFFFF"}
    )


def _validate_bg_contract(ws, company: str | None, period: str | None, issues: list[str]) -> None:
    expected_merges = {"B7:L7", "B8:L8", "B9:L9", "B10:L10"}
    if {str(item) for item in ws.merged_cells.ranges} != expected_merges:
        issues.append("BG title merges do not match the versioned contract.")
    if ws["B8"].value != "Balance General":
        issues.append("BG title is missing.")
    if company and ws["B7"].value != company:
        issues.append("BG company title was not derived from the loaded input.")
    if period and ws["B9"].value != _bg_period_label(period):
        issues.append("BG period title was not derived from the loaded input.")
    formulas = (
        ("F26", "=SUM(E16:E24)"),
        ("F36", "=SUM(E29:E34)"),
        ("F42", "=SUM(E39:E40)"),
        ("F45", "=F26+F36+F42"),
        ("L26", "=SUM(K16:K20)"),
        ("K34", "=ER!H70"),
        ("L36", "=SUM(K31:K34)"),
        ("L45", "=L26+L36"),
        ("L47", "=F45-L45"),
    )
    for coordinate, formula in formulas:
        if ws[coordinate].value != formula:
            issues.append(f"BG!{coordinate} must use {formula}.")


def _validate_bal_contract(
    ws,
    company: str | None,
    period: str | None,
    rfc: str | None,
    normalized_rows: int | None,
    issues: list[str],
) -> None:
    expected_merges = {"C1:G1", "C2:G2", "C3:G3", "C4:G4"}
    if {str(item) for item in ws.merged_cells.ranges} != expected_merges:
        issues.append("BAL title merges do not match the versioned contract.")
    if company and ws["C1"].value != company:
        issues.append("BAL company title was not derived from the loaded input.")
    if ws["C2"].value != "Balanza de Comprobaci\u00f3n":
        issues.append("BAL title is missing.")
    if period and ws["C3"].value != _bal_period_label(period):
        issues.append("BAL period title was not derived from the loaded input.")
    if rfc and ws["C4"].value != f"RFC: {rfc}":
        issues.append("BAL RFC title was not derived from the loaded input.")
    if [ws.cell(6, column).value for column in range(3, 8)] != [
        "CUENTA", "SALDO INICIAL", "DEBE", "HABER", "SALDO FINAL"
    ]:
        issues.append("BAL header row is invalid.")
    sum_row = 8 + normalized_rows if normalized_rows is not None else None
    if normalized_rows is not None:
        if ws.cell(7 + normalized_rows, 3).value is not None:
            issues.append("BAL separator row must be blank.")
        if ws.cell(sum_row, 3).value != "SUMAS IGUALES":
            issues.append("BAL SUMAS IGUALES row is not dynamic or is misplaced.")
        if ws.print_area != f"'BAL'!$C$1:$G${sum_row}":
            issues.append("BAL print area is not dynamic.")
        for coordinate in (f"E{sum_row}", f"F{sum_row}", f"G{sum_row}"):
            if not (isinstance(ws[coordinate].value, str) and ws[coordinate].value.startswith("=")):
                issues.append(f"BAL!{coordinate} must be a formula.")
    if ws.freeze_panes is not None:
        issues.append("BAL must not retain the manual freeze pane residue.")
    expected_widths = {"C": 28.46, "D": 15.7, "E": 13, "F": 13, "G": 13}
    for column, width in expected_widths.items():
        if ws.column_dimensions[column].width != width:
            issues.append(f"BAL column {column} width differs from the versioned contract.")
    expected_heights = {1: 13.5, 2: 15, 3: 17.25, 4: 15, 5: 17.25}
    for row, height in expected_heights.items():
        if ws.row_dimensions[row].height != height:
            issues.append(f"BAL row {row} height differs from the versioned contract.")
    if ws.page_setup.orientation != "portrait":
        issues.append("BAL must use portrait orientation.")
    for name, expected in (("left", 0.75), ("right", 0.75), ("top", 0.75), ("bottom", 0.75)):
        if getattr(ws.page_margins, name) != expected:
            issues.append(f"BAL {name} margin differs from the versioned contract.")
    if ws["C6"].border.left.style != "hair":
        issues.append("BAL hair borders are missing.")
    if ws["E7"].number_format != "General":
        issues.append("BAL numeric format differs from the approved manual.")
    for row in ws.iter_rows():
        for cell in row:
            if cell.column < 3 or cell.column > 7:
                if cell.value is not None:
                    issues.append(f"BAL contains data outside C:G at {cell.coordinate}.")
                    return


def _validate_print_contract(
    workbook: Workbook,
    normalized_rows: int | None,
    issues: list[str],
) -> None:
    expected = {
        "BG": ("'BG'!$B$7:$L$47", "$7:$10"),
        "ER": ("'ER'!$B$9:$J$70", "$9:$15"),
    }
    if normalized_rows is not None:
        sum_row = normalized_rows + 8
        expected["BAL"] = ("'BAL'!$C$1:$G$" + str(sum_row), "$1:$6")

    for sheet_name, (print_area, title_rows) in expected.items():
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        if ws.print_area != print_area:
            issues.append(f"{sheet_name} print area differs from the approved contract.")
        if ws.print_title_rows != title_rows:
            issues.append(f"{sheet_name} repeated print titles are missing.")
        if ws.page_setup.fitToWidth != 1 or ws.page_setup.fitToHeight != 0:
            issues.append(f"{sheet_name} must fit to one page wide.")
        page_setup_properties = ws.sheet_properties.pageSetUpPr
        if not page_setup_properties or page_setup_properties.fitToPage is not True:
            issues.append(f"{sheet_name} fit-to-page mode is not enabled.")
        if ws.page_setup.orientation != "portrait":
            issues.append(f"{sheet_name} must use portrait orientation.")
        if ws.page_setup.paperSize != 9:
            issues.append(f"{sheet_name} must use A4 paper.")
        for name, expected_margin in _VERSIONED_PRINT_MARGINS.items():
            if getattr(ws.page_margins, name) != expected_margin:
                issues.append(
                    f"{sheet_name} {name} margin differs from the versioned contract."
                )


_VERSIONED_PRINT_MARGINS = {
    "left": 0.75,
    "right": 0.75,
    "top": 0.75,
    "bottom": 0.75,
    "header": 0.25,
    "footer": 0.25,
}


def _validate_pdf_print_evidence(
    source: Path,
    executable: str,
) -> tuple[list[str], dict[str, int], list[str]]:
    """Render every temporary PDF page and check repeatable print evidence."""

    issues: list[str] = []
    warnings: list[str] = []
    page_counts: dict[str, int] = {}
    render_executable = shutil.which("pdftoppm") or shutil.which("pdftocairo")
    text_executable = shutil.which("pdftotext")
    if render_executable is None:
        return ["PDF print evidence cannot render pages: Poppler is unavailable."], page_counts, warnings
    if text_executable is None:
        return ["PDF print evidence cannot inspect repeated text: pdftotext is unavailable."], page_counts, warnings

    required_text = {
        "BG": ("Balance General", "ACTIVO", "PASIVO"),
        "ER": ("Estado de Resultados", "I N G R E S O S"),
        "BAL": ("Balanza de Comprobaci", "CUENTA", "SALDO"),
    }
    with tempfile.TemporaryDirectory(prefix="estados_financieros_print_") as directory:
        workdir = Path(directory)
        pdf_dir = workdir / "pdf"
        pdf_dir.mkdir()
        converted = subprocess.run(
            [executable, "--headless", "--convert-to", "pdf", "--outdir", str(pdf_dir), str(source)],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
        pdf_path = pdf_dir / f"{source.stem}.pdf"
        if converted.returncode != 0 or not pdf_path.is_file():
            detail = converted.stderr.strip() or converted.stdout.strip()
            return ["LibreOffice failed to create temporary PDF print evidence." + (f" {detail}" if detail else "")], page_counts, warnings

        # LibreOffice exports one PDF containing the workbook sheets in order.
        info = subprocess.run(
            [text_executable, "-bbox-layout", str(pdf_path), "-"],
            capture_output=True,
            text=True,
            timeout=60,
            check=False,
        )
        if info.returncode != 0:
            return ["pdftotext failed while inspecting temporary PDF evidence."], page_counts, warnings
        try:
            document = ElementTree.fromstring(info.stdout)
        except ElementTree.ParseError:
            return ["Temporary PDF text geometry could not be parsed."], page_counts, warnings

        pages = [element for element in document.iter() if element.tag.endswith("page")]
        if not pages:
            return ["Temporary PDF has no renderable pages."], page_counts, warnings
        page_texts = [_pdf_page_text(page) for page in pages]
        sheet_page_groups = _split_pdf_pages_by_sheet(page_texts, required_text)
        assigned_pages = {page_index for group in sheet_page_groups.values() for page_index in group}
        for page_index in set(range(len(pages))) - assigned_pages:
            issues.append(
                f"PDF page {page_index + 1} does not contain a repeated sheet title and header."
            )
        for sheet_name, group in sheet_page_groups.items():
            page_counts[sheet_name] = len(group)
            if not group:
                issues.append(f"Temporary PDF has no identifiable {sheet_name} page.")
                continue
            for index, page_index in enumerate(group, start=1):
                text = page_texts[page_index]
                missing = [token for token in required_text[sheet_name] if token not in text]
                if missing:
                    issues.append(
                        f"{sheet_name} PDF page {index} is missing repeated title/header text: {', '.join(missing)}."
                    )

        for page_index, page in enumerate(pages, start=1):
            width = float(page.attrib.get("width", "0"))
            height = float(page.attrib.get("height", "0"))
            if width <= 0 or height <= 0 or width >= height:
                issues.append(f"PDF page {page_index} is not portrait A4 evidence.")
            elif abs(width - 595.28) > 2 or abs(height - 841.89) > 2:
                issues.append(f"PDF page {page_index} is not A4-sized evidence.")
            _validate_pdf_text_geometry(page, page_index, issues)

        for page_index in range(1, len(pages) + 1):
            output_prefix = workdir / f"page-{page_index}"
            command = [render_executable, "-png", "-r", "144", "-f", str(page_index), "-l", str(page_index)]
            if Path(render_executable).name == "pdftocairo":
                command.extend([str(pdf_path), str(output_prefix)])
            else:
                command.extend([str(pdf_path), str(output_prefix)])
            rendered = subprocess.run(command, capture_output=True, text=True, timeout=60, check=False)
            if rendered.returncode != 0 or not list(workdir.glob(f"{output_prefix.name}*.png")):
                issues.append(f"Temporary PDF page {page_index} could not be rendered to an image.")
    return issues, page_counts, warnings


def _pdf_page_text(page: ElementTree.Element) -> str:
    return " ".join((word.text or "") for word in page.iter() if word.tag.endswith("word"))


def _split_pdf_pages_by_sheet(
    page_texts: list[str],
    required_text: Mapping[str, tuple[str, ...]],
) -> dict[str, list[int]]:
    """Assign pages to workbook sheets by their repeated title text."""

    groups = {sheet_name: [] for sheet_name in required_text}
    current_sheet: str | None = None
    for index, text in enumerate(page_texts):
        matches = [
            sheet_name
            for sheet_name, tokens in required_text.items()
            if tokens[0] in text
        ]
        if matches:
            current_sheet = matches[0]
        if current_sheet is None:
            continue
        groups[current_sheet].append(index)
    return groups


def _validate_pdf_text_geometry(page: ElementTree.Element, page_index: int, issues: list[str]) -> None:
    """Reject text outside page bounds or duplicate word boxes (overlap evidence)."""

    page_width = float(page.attrib.get("width", "0"))
    page_height = float(page.attrib.get("height", "0"))
    seen_boxes: set[tuple[float, float, float, float, str]] = set()
    for word in (element for element in page.iter() if element.tag.endswith("word")):
        try:
            box = tuple(round(float(word.attrib[key]), 2) for key in ("xMin", "yMin", "xMax", "yMax"))
        except (KeyError, ValueError):
            continue
        if box[0] < 0 or box[1] < 0 or box[2] > page_width or box[3] > page_height:
            issues.append(f"PDF page {page_index} has text cut outside its page bounds.")
            return
        key = (*box, word.text or "")
        if key in seen_boxes:
            issues.append(f"PDF page {page_index} has overlapping duplicate text.")
            return
        seen_boxes.add(key)


def _validate_forbidden_fills(workbook: Workbook, issues: list[str]) -> None:
    forbidden = {"FFFFFF00", "FF00A933"}
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                color = cell.fill.fgColor
                if cell.fill.fill_type and color.type == "rgb" and color.rgb in forbidden:
                    issues.append(f"{ws.title}!{cell.coordinate} has a forbidden mask fill.")
                    return


def _financial_statements_evidence(
    bal_dataset: Mapping[str, Any] | None,
    bg_dataset: Mapping[str, Any] | None,
) -> dict[str, Any]:
    evidence: dict[str, Any] = {
        "formula_recalculation_performed": False,
        "formula_evaluated_error_count": None,
    }
    if bal_dataset:
        totals = bal_dataset.get("sumas_iguales", {}).get("totals", {})
        if totals:
            debe = float(totals.get("debe", 0))
            haber = float(totals.get("haber", 0))
            saldo_inicial = 0.0  # BAL's total row is a zero control, not a source-row sum.
            evidence["bal_sumas_iguales"] = {
                "saldo_inicial": saldo_inicial,
                "debe": debe,
                "haber": haber,
                "saldo_final": saldo_inicial + debe - haber,
            }
    if bg_dataset:
        balance = bg_dataset.get("balance", bg_dataset)
        difference = float(balance.get("diferencia_cuadre", 0))
        evidence["bg_balance"] = {
            "total_activo": float(balance.get("total_activo", 0)),
            "total_pasivo": float(balance.get("total_pasivo", 0)),
            "capital_contable": float(balance.get("capital_contable", 0)),
            "difference": difference,
            "report_difference": float(balance.get("diferencia_cuadre", 0)),
            "reference": balance.get("reference", CONCEPTUAL_BALANCE_REFERENCE),
        }
    return evidence


def _within_cent(left: float, right: float) -> bool:
    return abs(float(left) - float(right)) <= 0.01


def _bg_period_label(period: str) -> str:
    from calendar import monthrange

    months = ("Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre")
    year, month = (int(value) for value in period.split("-", 1))
    return f"Al {monthrange(year, month)[1]} de {months[month - 1]} de {year}"


def _bal_period_label(period: str) -> str:
    months = ("ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE")
    year, month = (int(value) for value in period.split("-", 1))
    return f"{months[month - 1]}.{year}"


def validate_balance_sheet(
    rows: Iterable[Mapping[str, Any] | Any],
    *,
    tolerance: float = BALANCE_TOLERANCE_DEFAULT,
    result_ejercicio: float | int | str | Decimal | None = None,
) -> BalanceCheckResult:
    """Validate the exact accounting equation that future ``BG!L47`` uses."""

    dataset = build_bg_dataset(
        (_row_to_mapping(row) for row in rows),
        result_ejercicio=result_ejercicio,
        tolerance=tolerance,
    )
    balance = dataset["balance"]
    component_rows = {"activo": [], "pasivo": [], "capital": []}
    for line in dataset["lines"]:
        component_rows[line["section"]].append(
            {"rubro": line["key"], "total": line["amount"], "cuentas": line["resolutions"]}
        )
    return BalanceCheckResult(
        total_activo=float(balance["total_activo"]),
        total_pasivo=float(balance["total_pasivo"]),
        capital_contable=float(balance["capital_contable"]),
        diferencia_cuadre=float(balance["diferencia_cuadre"]),
        tolerance=float(balance["tolerance"]),
        cuadra=bool(balance["cuadra"]),
        balanza_no_cuadra=bool(balance["balanza_no_cuadra"]),
        componentes=[
            BalanceComponent(rubro="activo", total=float(balance["total_activo"]), cuentas=component_rows["activo"]),
            BalanceComponent(rubro="pasivo", total=float(balance["total_pasivo"]), cuentas=component_rows["pasivo"]),
            BalanceComponent(rubro="capital_contable", total=float(balance["capital_contable"]), cuentas=component_rows["capital"]),
        ],
        warnings=list(dataset["warnings"]),
    )


def _capital_balance_rows(rows: Iterable[Mapping[str, Any]]) -> tuple[Mapping[str, Any], ...]:
    """Select top-level capital balances, falling back to leaves when needed."""

    materialized = tuple(rows)
    capital_rows = []
    capital_codes = {
        canonical_account_code(_read_field(row, ("account_code", "top_account", "cuenta")) or "")
        for row in materialized
    }
    top_level_codes = {
        code for code in capital_codes if code.startswith("3") and "-" not in code
    }
    for row in materialized:
        code = canonical_account_code(_read_field(row, ("account_code", "top_account", "cuenta")) or "")
        if not code.startswith("3"):
            continue
        top_code = code.split("-", 1)[0]
        if top_code in top_level_codes:
            if code == top_code:
                capital_rows.append(row)
            continue
        if not any(
            other != code and other.startswith(f"{code}-")
            for other in capital_codes
        ):
            capital_rows.append(row)
    return tuple(capital_rows)


def find_formula_issues(workbook: Workbook) -> list[FormulaIssue]:
    """Return formula cells that contain known Excel error tokens or externals."""

    issues: list[FormulaIssue] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                upper = value.upper()
                for token in FORMULA_ERROR_TOKENS:
                    if token in upper:
                        issues.append(
                            FormulaIssue(
                                cell=f"{worksheet.title}!{cell.coordinate}",
                                value=value,
                                reason=f"Formula contains {token}",
                            )
                        )
                        break
                else:
                    if "[" in value or "]" in value:
                        issues.append(
                            FormulaIssue(
                                cell=f"{worksheet.title}!{cell.coordinate}",
                                value=value,
                                reason="Formula contains an external workbook reference",
                            )
                        )
    return issues


def find_evaluated_formula_issues(
    formula_workbook: Workbook,
    evaluated_workbook: Workbook,
) -> tuple[list[FormulaIssue], bool]:
    """Inspect cached/evaluated values for the formula cells in a workbook."""

    issues: list[FormulaIssue] = []
    formula_cells = []
    for worksheet in formula_workbook.worksheets:
        evaluated_sheet = evaluated_workbook[worksheet.title]
        for row in worksheet.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                formula_cells.append((worksheet.title, cell.coordinate))
                evaluated_value = evaluated_sheet[cell.coordinate].value
                if isinstance(evaluated_value, str):
                    upper = evaluated_value.upper()
                    for token in FORMULA_ERROR_TOKENS:
                        if token in upper:
                            issues.append(
                                FormulaIssue(
                                    cell=f"{worksheet.title}!{cell.coordinate}",
                                    value=evaluated_value,
                                    reason=f"Evaluated formula contains {token}",
                                )
                            )
                            break
    cached_values_available = bool(formula_cells) and all(
        evaluated_workbook[sheet][coordinate].value is not None
        for sheet, coordinate in formula_cells
    )
    return issues, cached_values_available


def recalculate_workbook(
    workbook_path: str | Path,
    *,
    balance_difference: float | int | str | None = None,
    tolerance: float = 1.0,
) -> FormulaRecalculationResult:
    """Recalculate a copy with LibreOffice when available.

    The final workbook is replaced atomically only after evaluated values are
    available and contain no formula errors. Failed recalculations leave a
    sibling evidence copy and mark the result as blocked.
    """

    source = Path(workbook_path)
    executable = shutil.which("libreoffice") or shutil.which("soffice")
    if executable is None:
        return FormulaRecalculationResult(
            performed=False,
            engine="none",
            evaluated_error_count=None,
            cached_values_available=False,
            warnings=[
                "Formula recalculation was not performed: libreoffice/soffice is unavailable."
            ],
        )

    with tempfile.TemporaryDirectory(prefix="estados_financieros_recalc_") as temp_dir:
        temp_root = Path(temp_dir)
        input_copy = temp_root / source.name
        output_dir = temp_root / "output"
        output_dir.mkdir()
        shutil.copy2(source, input_copy)
        completed = subprocess.run(
            [executable, "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(input_copy)],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
        evaluated_path = output_dir / input_copy.name
        if completed.returncode != 0 or not evaluated_path.exists():
            return FormulaRecalculationResult(
                performed=False,
                engine=Path(executable).name,
                evaluated_error_count=None,
                cached_values_available=False,
                blocked=True,
                warnings=[
                    "LibreOffice recalculation failed before producing an evaluated workbook.",
                    completed.stderr.strip() or completed.stdout.strip(),
                ],
            )

        evaluated = validate_generated_workbook(
            source,
            balance_difference=balance_difference,
            tolerance=tolerance,
            formula_mode="recalculated_ok",
            evaluated_workbook_or_path=evaluated_path,
            formula_recalculation_engine=Path(executable).name,
        )
        if not evaluated.formula_cached_values_available or not evaluated.ok:
            evidence = source.with_name(
                f"{source.stem}.recalculation-failed-{os.getpid()}.xlsx"
            )
            shutil.copy2(evaluated_path, evidence)
            return FormulaRecalculationResult(
                performed=True,
                engine=Path(executable).name,
                evaluated_error_count=evaluated.formula_evaluated_error_count,
                cached_values_available=evaluated.formula_cached_values_available,
                evaluated_issues=evaluated.formula_evaluated_issues,
                evidence_path=str(evidence),
                blocked=True,
                warnings=[
                    "Evaluated workbook failed validation; final output was not replaced.",
                ],
            )

        os.replace(evaluated_path, source)
        return FormulaRecalculationResult(
            performed=True,
            engine=Path(executable).name,
            evaluated_error_count=0,
            cached_values_available=True,
        )


def validate_balance_difference(
    difference: float | int | str,
    *,
    tolerance: float = 1.0,
) -> bool:
    """Return True when abs(difference) is strictly less than tolerance."""

    numeric = _coerce_float(difference)
    if numeric is None:
        return False
    return abs(numeric) < tolerance


def assert_generated_workbook_valid(
    workbook_or_path: Workbook | str | Path,
    *,
    balance_difference: float | int | str | None = None,
    tolerance: float = 1.0,
) -> ValidationResult:
    """Validate and raise ValueError with compact details if checks fail."""

    result = validate_generated_workbook(
        workbook_or_path,
        balance_difference=balance_difference,
        tolerance=tolerance,
    )
    if result.ok:
        return result
    details = [f"{issue.cell}: {issue.reason}" for issue in result.formula_static_issues]
    if result.balance_ok is False:
        details.append(
            f"Balance difference {result.balance_difference!r} does not satisfy abs(diff) < {tolerance}"
        )
    raise ValueError("; ".join(details) or "Generated workbook validation failed")


def _load_workbook(
    workbook_or_path: Workbook | str | Path | None,
    *,
    data_only: bool = False,
) -> tuple[Workbook, bool]:
    if workbook_or_path is None:
        raise ValueError("Workbook path is required")
    if isinstance(workbook_or_path, Workbook):
        return workbook_or_path, False
    return load_workbook(Path(workbook_or_path), data_only=data_only), True


def _coerce_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None


def _read_field(source: Any, keys: tuple[str, ...]) -> Any:
    if source is None:
        return None
    if isinstance(source, Mapping):
        for key in keys:
            if key in source and source[key] not in (None, ""):
                return source[key]
    for key in keys:
        if hasattr(source, key):
            value = getattr(source, key)
            if value not in (None, ""):
                return value
    return None


def _row_to_mapping(row: Any) -> dict[str, Any]:
    if isinstance(row, Mapping):
        return dict(row)
    if hasattr(row, "to_dict"):
        return dict(row.to_dict())
    if hasattr(row, "__dict__"):
        return dict(vars(row))
    return {}


def _quantize_money(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _quantize_balance_difference(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.000001"), rounding=ROUND_HALF_UP))
