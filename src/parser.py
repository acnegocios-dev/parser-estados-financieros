from __future__ import annotations

import re
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass, field, replace
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

try:
    from .period import PeriodVariables, extract_period_variables
except ImportError:  # pragma: no cover - supports PYTHONPATH=src usage.
    from period import PeriodVariables, extract_period_variables


REQUIRED_SHEET = "Balanza"
REQUIRED_COLUMNS = ("Cuenta", "Saldo Inicial", "Debe", "Haber", "SaldoFinal")
OOXML_ZIP = "ooxml_zip"
OLE_XLS = "ole_xls"
HTML_EXCEL = "html_excel"
UNKNOWN_FILE = "unknown"

_ACCOUNT_RE = re.compile(
    r"^\s*(?P<code>\d+(?:-\s*[A-Za-z0-9]+)*)\s+(?P<name>.+?)\s*$"
)


@dataclass(frozen=True)
class BalanzaRow:
    source_row: int
    account_raw: str
    account_code: str
    account_name: str
    top_account: str
    saldo_inicial: Decimal
    debe: Decimal
    haber: Decimal
    saldo_final: Decimal
    # These fields are evidence from the private company catalog.  They are
    # intentionally optional because a balanza can be parsed before its
    # catalog is available; no value is inferred from the account prefix.
    parent_code: str | None = None
    nature: str | None = None
    sat_group_code: str | None = None
    catalog_match: str = "not_provided"

    def to_dict(self) -> dict[str, object]:
        data = asdict(self)
        for key in ("saldo_inicial", "debe", "haber", "saldo_final"):
            data[key] = str(data[key])
        return data


@dataclass(frozen=True)
class RowIssue:
    source_row: int
    message: str
    values: tuple[Any, ...] = field(default_factory=tuple)

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class ParserWarning:
    code: str
    message: str
    source_row: int | None = None

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class ParsedBalanza:
    source_path: str
    detected_mime_kind: str
    sheet_name: str
    period: PeriodVariables
    company_name: str | None
    content_rfc: str | None
    content_period_ym: str | None
    header_row: int
    rows: tuple[BalanzaRow, ...]
    empty_rows: tuple[int, ...]
    structure_issues: tuple[RowIssue, ...]
    warnings: tuple[ParserWarning, ...]

    def to_dict(self) -> dict[str, object]:
        return {
            "source_path": self.source_path,
            "detected_mime_kind": self.detected_mime_kind,
            "sheet_name": self.sheet_name,
            "period": self.period.to_dict(),
            "company_name": self.company_name,
            "content_rfc": self.content_rfc,
            "content_period_ym": self.content_period_ym,
            "header_row": self.header_row,
            "rows": [row.to_dict() for row in self.rows],
            "empty_rows": list(self.empty_rows),
            "structure_issues": [issue.to_dict() for issue in self.structure_issues],
            "warnings": [warning.to_dict() for warning in self.warnings],
        }


def load_ooxml_workbook(path: str | Path):
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    detected_mime_kind = detect_mime_kind(workbook_path)
    if detected_mime_kind != OOXML_ZIP:
        raise ValueError(
            f"Workbook is not an OOXML ZIP package: {workbook_path} "
            f"(detected_mime_kind={detected_mime_kind})"
        )

    with workbook_path.open("rb") as fh:
        return load_workbook(fh, data_only=True, read_only=False)


def detect_mime_kind(path: str | Path) -> str:
    """Detect workbook container by file content, not by extension."""

    workbook_path = Path(path)
    if zipfile.is_zipfile(workbook_path):
        return OOXML_ZIP

    with workbook_path.open("rb") as fh:
        prefix = fh.read(512)

    if prefix.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return OLE_XLS

    stripped = prefix.lstrip().lower()
    if stripped.startswith((b"<!doctype html", b"<html", b"<?xml")):
        return HTML_EXCEL

    return UNKNOWN_FILE


def parse_balanza(
    path: str | Path,
    sheet_name: str = REQUIRED_SHEET,
    *,
    catalog_rows: Iterable[Any] | None = None,
) -> ParsedBalanza:
    workbook_path = Path(path)
    period = extract_period_variables(workbook_path)
    detected_mime_kind = detect_mime_kind(workbook_path)
    if detected_mime_kind != OOXML_ZIP:
        raise ValueError(
            f"Unsupported workbook content kind: {detected_mime_kind}. "
            "Only OOXML ZIP workbooks are supported in the initial Auditalo parser."
        )
    workbook = load_ooxml_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Required sheet '{sheet_name}' not found. Sheets: {workbook.sheetnames}")

    worksheet = workbook[sheet_name]
    company_name, content_rfc, content_period_ym = _extract_sheet_metadata(worksheet)
    _validate_sheet_metadata(period, content_rfc, content_period_ym)
    header_row, column_map = _find_header_row(worksheet.iter_rows(values_only=True))

    rows: list[BalanzaRow] = []
    empty_rows: list[int] = []
    issues: list[RowIssue] = []
    warnings: list[ParserWarning] = []
    if content_rfc is None:
        warnings.append(
            ParserWarning(
                code="rfc_no_encontrado_en_contenido",
                message="No se encontro RFC en el contenido; se conserva el RFC del nombre de archivo.",
            )
        )
    if content_period_ym is None:
        warnings.append(
            ParserWarning(
                code="periodo_no_encontrado_en_contenido",
                message="No se encontro periodo interno; se conserva el periodo del nombre de archivo.",
            )
        )
    data_started = False

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        values = tuple(
            worksheet.cell(row=row_number, column=column_map[column]).value
            for column in REQUIRED_COLUMNS
        )
        if _is_empty(values):
            empty_rows.append(row_number)
            continue

        account_raw = _clean(values[0])
        parsed_account = _parse_account(account_raw)
        if parsed_account is None:
            if data_started:
                if _is_repeated_header(values):
                    warnings.append(
                        ParserWarning(
                            code="encabezado_repetido_ignorado",
                            message="Se ignoro un encabezado repetido dentro de los datos.",
                            source_row=row_number,
                        )
                    )
                    continue
                issues.append(RowIssue(row_number, "Missing or invalid account code.", values))
            continue

        data_started = True
        amounts = _parse_amounts(values[1:])
        if isinstance(amounts, RowIssue):
            issues.append(RowIssue(row_number, amounts.message, values))
            continue

        account_code, account_name = parsed_account
        rows.append(
            BalanzaRow(
                source_row=row_number,
                account_raw=account_raw,
                account_code=account_code,
                account_name=account_name,
                top_account=account_code.split("-", 1)[0],
                saldo_inicial=amounts[0],
                debe=amounts[1],
                haber=amounts[2],
                saldo_final=amounts[3],
            )
        )

    duplicate_counts = Counter(row.account_code for row in rows)
    for account_code, count in sorted(duplicate_counts.items()):
        if count > 1:
            warnings.append(
                ParserWarning(
                    code="cuenta_repetida_agregada",
                    message=(
                        f"La cuenta {account_code} aparece {count} veces; se conservaran y agregaran "
                        "sus renglones de detalle."
                    ),
                )
            )

    if not rows:
        issues.append(RowIssue(header_row, "No account rows were parsed.", ()))
    if catalog_rows is not None:
        rows = list(enrich_balanza_rows(rows, catalog_rows))

    return ParsedBalanza(
        source_path=str(workbook_path),
        detected_mime_kind=detected_mime_kind,
        sheet_name=sheet_name,
        period=period,
        company_name=company_name,
        content_rfc=content_rfc,
        content_period_ym=content_period_ym,
        header_row=header_row,
        rows=tuple(rows),
        empty_rows=tuple(empty_rows),
        structure_issues=tuple(issues),
        warnings=tuple(warnings),
    )


def parse_balanza_dict(path: str | Path, sheet_name: str = REQUIRED_SHEET) -> dict[str, object]:
    return parse_balanza(path, sheet_name=sheet_name).to_dict()


def enrich_balanza_rows(
    rows: Iterable[BalanzaRow], catalog_rows: Iterable[Any]
) -> tuple[BalanzaRow, ...]:
    """Attach catalog evidence by exact text code without repairing either source.

    A duplicate catalog key remains explicitly ambiguous.  The parser never
    guesses a parent, nature, or SAT group from a first digit or account name.
    """

    by_code: dict[str, list[Any]] = {}
    for catalog_row in catalog_rows:
        code = _canonical_catalog_code(getattr(catalog_row, "account_code", None))
        if code:
            by_code.setdefault(code, []).append(catalog_row)

    enriched: list[BalanzaRow] = []
    for row in rows:
        candidates = by_code.get(_canonical_catalog_code(row.account_code), [])
        if len(candidates) == 1:
            catalog = candidates[0]
            enriched.append(
                replace(
                    row,
                    parent_code=_clean(getattr(catalog, "parent_code", None)) or None,
                    nature=_clean(getattr(catalog, "nature", None)) or None,
                    sat_group_code=_clean(getattr(catalog, "sat_group_code", None)) or None,
                    catalog_match="matched",
                )
            )
        elif candidates:
            enriched.append(replace(row, catalog_match="ambiguous"))
        else:
            enriched.append(replace(row, catalog_match="missing"))
    return tuple(enriched)


def _canonical_catalog_code(value: Any) -> str:
    text = _clean(value).upper()
    return "-".join(part for part in re.split(r"[-.\s]+", text) if part)


def _find_header_row(rows: Iterable[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    required = {_normalize_header(column): column for column in REQUIRED_COLUMNS}
    for row_number, row in enumerate(rows, start=1):
        headers = {_normalize_header(value): index for index, value in enumerate(row, start=1)}
        if all(normalized in headers for normalized in required):
            return row_number, {
                original: headers[normalized] for normalized, original in required.items()
            }
    raise ValueError(f"Required columns not found: {', '.join(REQUIRED_COLUMNS)}")


def _parse_account(value: str) -> tuple[str, str] | None:
    match = _ACCOUNT_RE.match(value)
    if not match:
        return None
    account_code = re.sub(r"\s+", "", match.group("code"))
    return account_code, match.group("name").strip()


def _parse_amounts(values: tuple[Any, ...]) -> tuple[Decimal, Decimal, Decimal, Decimal] | RowIssue:
    parsed: list[Decimal] = []
    for value in values:
        try:
            parsed.append(_to_decimal(value))
        except (InvalidOperation, TypeError, ValueError):
            return RowIssue(0, f"Invalid numeric value: {value!r}", values)
    return parsed[0], parsed[1], parsed[2], parsed[3]


def _to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value).replace(",", "").strip())


def _is_empty(values: tuple[Any, ...]) -> bool:
    return all(value is None or _clean(value) == "" for value in values)


def _is_repeated_header(values: tuple[Any, ...]) -> bool:
    normalized = {_normalize_header(value) for value in values}
    required = {_normalize_header(column) for column in REQUIRED_COLUMNS}
    return required.issubset(normalized)


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", _clean(value).casefold())


def _validate_sheet_metadata(
    period: PeriodVariables,
    content_rfc: str | None,
    content_period_ym: str | None,
) -> None:
    if content_period_ym and content_period_ym != period.period_ym:
        raise ValueError(
            f"Filename period {period.period_ym} differs from sheet period {content_period_ym}."
        )
    if content_rfc and content_rfc.upper() != period.rfc:
        raise ValueError(f"Filename RFC {period.rfc} differs from sheet RFC {content_rfc.upper()}.")


def _extract_sheet_metadata(worksheet) -> tuple[str | None, str | None, str | None]:
    company_name: str | None = None
    content_rfc: str | None = None
    content_period_ym: str | None = None
    rfc_pattern = re.compile(r"\b[A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{3}\b")
    period_pattern = re.compile(r"(\d{4})[-/](\d{1,2})")

    for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 12), values_only=True):
        for value in row:
            text = _clean(value)
            if not text:
                continue
            rfc_match = rfc_pattern.search(text)
            if rfc_match and "periodo" not in text.casefold():
                if content_rfc is None:
                    content_rfc = rfc_match.group(0).upper()
                if company_name is None:
                    company_name = rfc_pattern.sub("", text).strip()
            elif company_name is None and "periodo" not in text.casefold():
                normalized = _normalize_header(text)
                if normalized not in {_normalize_header(column) for column in REQUIRED_COLUMNS}:
                    company_name = text
            if content_period_ym is None:
                match = period_pattern.search(text)
                if match:
                    content_period_ym = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
            if company_name and content_rfc and content_period_ym:
                return company_name, content_rfc, content_period_ym

    return company_name, content_rfc, content_period_ym
