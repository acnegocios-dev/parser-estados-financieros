"""Self-contained BG, ER and BAL workbook generation for the financial module."""

from __future__ import annotations

import calendar
from dataclasses import dataclass, field, is_dataclass
from datetime import datetime, timezone
import json
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "sample-outputs"
ER_STYLE_SPEC_PATH = Path(__file__).with_name("er_style_spec.json")
BG_STYLE_SPEC_PATH = Path(__file__).with_name("bg_style_spec.json")
BAL_STYLE_SPEC_PATH = Path(__file__).with_name("bal_style_spec.json")

ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
MONTHS_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


@dataclass
class WorkbookBuildResult:
    workbook: Workbook
    output_path: Path | None
    company: str
    period: str
    source_path: str | None
    formula_cells: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    missing_accounts: list[str] = field(default_factory=list)
    generated_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))


ER_LAYOUT = [
    {"row": 17, "label": "I N G R E S O S", "kind": "section"},
    {"row": 18, "label": "Ingresos por servicios", "key": "ingresos_por_servicios"},
    {"row": 19, "label": "Descuentos o bonificaciones", "key": "descuentos_o_bonificaciones"},
    {"row": 20, "label": "INGRESOS NETOS", "kind": "subtotal", "formula": "SUM({col}18:{col}19)"},
    {"row": 22, "label": "C O S T O S", "kind": "section"},
    {"row": 23, "label": "Costo de Ventas", "key": "costo_de_ventas"},
    {"row": 25, "label": "UTILIDAD BRUTA", "kind": "subtotal", "formula": "{col}20-{col}23"},
    {"row": 27, "label": "Gastos de Operacion", "kind": "section"},
    {"row": 28, "label": "Sueldos y Salarios", "key": "sueldos_y_salarios"},
    {"row": 29, "label": "Impuestos y Derechos", "key": "impuestos_y_derechos"},
    {"row": 30, "label": "Honorarios", "key": "honorarios"},
    {"row": 31, "label": "Arrendamiento", "key": "arrendamiento"},
    {"row": 32, "label": "Seguros y Fianzas", "key": "seguros_y_fianzas"},
    {"row": 33, "label": "Servicios", "key": "servicios"},
    {"row": 34, "label": "Capacitacion al Personal", "key": "capacitacion_al_personal"},
    {"row": 35, "label": "Fletes y/o Mensajeria", "key": "fletes_y_o_mensajeria"},
    {"row": 36, "label": "Seguridad e higiene", "key": "seguridad_e_higiene"},
    {"row": 37, "label": "Mantenimiento", "key": "mantenimiento"},
    {"row": 38, "label": "Combustibles", "key": "combustibles"},
    {"row": 39, "label": "Propaganda y Publicidad", "key": "propaganda_y_publicidad"},
    {"row": 40, "label": "Cuotas y Suscripciones", "key": "cuotas_y_suscripciones"},
    {"row": 41, "label": "Gastos de Viaje", "key": "gastos_de_viaje"},
    {"row": 42, "label": "Herrajes y Herramientas", "key": "herrajes_y_herramientas"},
    {"row": 43, "label": "Papeleria y Art. de Oficina", "key": "papeleria_y_art_de_oficina"},
    {"row": 44, "label": "Depreciaciones", "key": "depreciaciones"},
    {"row": 45, "label": "Recargos", "key": "recargos"},
    {"row": 46, "label": "Varios", "key": "varios"},
    {"row": 47, "label": "Uniformes", "key": "uniformes"},
    {"row": 50, "label": "No deducibles", "key": "no_deducibles"},
    {"row": 51, "label": "GASTOS DE OPERACION", "kind": "subtotal", "formula": "SUM({col}28:{col}50)"},
    {"row": 53, "label": "UTILIDAD O (PERDIDA) DE OPERACION", "kind": "subtotal", "formula": "{col}25-{col}51"},
    {"row": 55, "label": "OTROS INGRESOS Y GASTOS", "kind": "section"},
    {"row": 56, "label": "Otros Productos", "key": "otros_productos"},
    {"row": 57, "label": "Otros Gastos", "key": "otros_gastos"},
    {"row": 58, "label": "TOTAL OTROS INGRESOS", "kind": "subtotal", "formula": "SUM({col}56:{col}57)"},
    {"row": 60, "label": "RES. INT. DE FINANCIAMIENTO", "kind": "section"},
    {"row": 61, "label": "Productos Financieros", "key": "productos_financieros"},
    {"row": 62, "label": "Gastos Financieros", "key": "gastos_financieros"},
    {"row": 63, "label": "TOTAL R. I. F.", "kind": "subtotal", "formula": "SUM({col}61:{col}62)"},
    {"row": 65, "label": "RESULTADO ANTES DE IMPUESTOS", "kind": "subtotal", "formula": "{col}53+{col}58+{col}63"},
    {"row": 67, "label": "ISR DEL EJERCICIO", "key": "isr_del_ejercicio"},
    {"row": 68, "label": "PTU DEL EJERCICIO", "key": "ptu_del_ejercicio"},
    {"row": 70, "label": "RESULTADO DEL EJERCICIO", "kind": "subtotal", "formula": "{col}65-{col}67-{col}68"},
]


def build_er_workbook(
    dataset: Any | None = None,
    *,
    metadata: Any | None = None,
    source_path: str | Path | None = None,
) -> WorkbookBuildResult:
    """Compatibility wrapper for the full financial-statements workbook."""

    bg_dataset, bal_dataset, metadata = _derive_supporting_datasets(
        dataset, metadata=metadata, source_path=source_path
    )
    return build_financial_statements_workbook(
        dataset,
        bg_dataset=bg_dataset,
        bal_dataset=bal_dataset,
        metadata=metadata,
        source_path=source_path,
    )


def build_financial_statements_workbook(
    er_dataset: Any | None = None,
    *,
    bg_dataset: Any | None = None,
    bal_dataset: Any | None = None,
    metadata: Any | None = None,
    source_path: str | Path | None = None,
) -> WorkbookBuildResult:
    """Build the self-contained, ordered ``BG`` + ``ER`` + ``BAL`` workbook."""

    metadata_values = _metadata_from_inputs(er_dataset, metadata, source_path)
    company = metadata_values["company"]
    period = metadata_values["period"]
    warnings = list(metadata_values["warnings"])
    missing_accounts = _coerce_list(_read_field(er_dataset, ("missing_accounts", "cuentas_no_encontradas")))
    missing_accounts.extend(_coerce_list(_read_field(er_dataset, ("missing", "not_found"))))

    wb = Workbook()
    _configure_recalculation_on_open(wb)
    bg = wb.active
    bg.title = "BG"
    er = wb.create_sheet("ER")
    bal = wb.create_sheet("BAL")
    for worksheet in wb.worksheets:
        worksheet.sheet_state = "visible"

    formula_cells: list[str] = []
    _write_bg_sheet(bg, company, period, bg_dataset or {}, formula_cells)
    _write_header(er, company, period)
    line_values = _index_dataset_lines(er_dataset, warnings)
    formula_cells.extend(_write_er_body(er, line_values, warnings))
    _style_er_sheet(er)
    _write_bal_sheet(bal, company, period, bal_dataset or {}, metadata, formula_cells)

    return WorkbookBuildResult(
        workbook=wb,
        output_path=None,
        company=company,
        period=period,
        source_path=str(source_path) if source_path is not None else metadata_values["source_path"],
        formula_cells=formula_cells,
        warnings=warnings,
        missing_accounts=missing_accounts,
    )


def _derive_supporting_datasets(
    er_dataset: Any,
    *,
    metadata: Any | None,
    source_path: str | Path | None,
) -> tuple[dict[str, Any], dict[str, Any], Any | None]:
    """Derive BG/BAL only from the loaded input, never from a manual workbook."""

    if source_path is None:
        return {}, {}, metadata
    try:
        from .engine import build_bal_dataset, build_bg_dataset, build_er_dataset
        from .parser import parse_balanza
    except ImportError:  # pragma: no cover - supports direct execution.
        from engine import build_bal_dataset, build_bg_dataset, build_er_dataset
        from parser import parse_balanza

    parsed = parse_balanza(source_path)
    effective_er = er_dataset or build_er_dataset(
        parsed.rows,
        company=parsed.company_name,
        period=parsed.period.period_ym,
        source_path=parsed.source_path,
    )
    merged_metadata = dict(_to_dict(metadata))
    merged_metadata.update({
        "company": parsed.company_name,
        "period": parsed.period.period_ym,
        "rfc": parsed.period.rfc,
        "source_path": parsed.source_path,
    })
    raw_amounts = _read_field(effective_er, ("raw_amounts",)) or {}
    return (
        build_bg_dataset(parsed.rows, result_ejercicio=raw_amounts.get("resultado_ejercicio")),
        build_bal_dataset(parsed.rows),
        merged_metadata,
    )


def _write_bg_sheet(
    ws,
    company: str,
    period: str,
    dataset: Any,
    formula_cells: list[str],
) -> None:
    spec = _load_style_spec(BG_STYLE_SPEC_PATH)
    ws["B7"] = company
    ws["B8"] = "Balance General"
    ws["B9"] = _bg_period_caption(period)
    ws["B10"] = "(Importes expresados en pesos)"

    ws["C13"] = "  A C T I V O"
    ws["C14"] = "C I R C U L A N T E"
    ws["I13"] = "  P A S I V O"
    ws["I14"] = "C I R C U L A N T E"
    ws["C28"] = "NO CIRCULANTE"
    ws["I29"] = "C A P I T A L"
    ws["I30"] = "C O N T A B L E"
    ws["C38"] = "OTROS ACTIVOS"

    lines = {
        str(line.get("key")): line
        for line in _coerce_list_of_dicts(_read_field(dataset, ("lines",)))
    }
    asset_layout = (
        ("caja_chica", "Caja Chica", "B16", "E16"),
        ("bancos", "Bancos", "B17", "E17"),
        ("inversiones_y_valores", "Inversiones y Valores", "B18", "E18"),
        ("cuentas_por_cobrar", "Cuentas por Cobrar", "B19", "E19"),
        ("contribuciones", "Contribuciones", "B20", "E20"),
        ("inventarios", "Inventarios", "B21", "E21"),
        ("deudores_diversos", "Deudores Diversos", "B22", "E22"),
        ("anticipo_a_proveedores", "Anticipo a Proveedores", "B23", "E23"),
        ("pagos_anticipados", "Pagos anticipados", "B24", "E24"),
        (None, "Herramientas", "B29", "E29"),
        ("mobiliario_y_equipo", "Mobiliario y Equipo", "B30", "E30"),
        ("equipo_de_computo", "Equipo de C\u00f3mputo", "B31", "E31"),
        ("equipo_de_transporte", "Equipo de transporte", "B32", "E32"),
        ("maquinaria_y_equipo", "Maquinaria y Equipo", "B33", "E33"),
        ("depreciacion_acumulada", "Depreciacion acumulada de activos", "B34", "E34"),
        (None, "Dep\u00f3sitos en Garant\u00eda", "B39", "E39"),
        (None, "Pagos Anticipados", "B40", "E40"),
    )
    liability_layout = (
        ("proveedores", "Proveedores", "H16", "K16"),
        ("acreedores_diversos", "Acreedores Diversos", "H17", "K17"),
        ("anticipos_de_clientes", "Anticipos de Clientes", "H18", "K18"),
        ("impuestos_por_pagar", "Impuestos por Pagar", "H19", "K19"),
        ("otros_pasivos_ptu", "Otros Pasivos PTU", "H20", "K20"),
        ("capital_social", "Capital Social", "H31", "K31"),
        (
            "aportaciones_para_aumentos_de_capital",
            "Aportaciones para aumentos de Capital",
            "H32",
            "K32",
        ),
        (
            "resultados_de_ejercicios_anteriores",
            "Resultados de Ejercicios Ant.",
            "H33",
            "K33",
        ),
    )
    for key, label, label_cell, amount_cell in asset_layout + liability_layout:
        ws[label_cell] = label
        ws[amount_cell] = lines.get(key, {}).get("amount", 0) if key else 0

    ws["B26"] = "TOTAL CIRCULANTE"
    ws["F26"] = "=SUM(E16:E24)"
    ws["B36"] = "TOTAL NO CIRCULANTE"
    ws["F36"] = "=SUM(E29:E34)"
    ws["B42"] = "TOTAL OTROS ACTIVOS"
    ws["F42"] = "=SUM(E39:E40)"
    ws["B45"] = "TOTAL DEL ACTIVO"
    ws["F45"] = "=F26+F36+F42"

    ws["H26"] = "TOTAL PASIVO"
    ws["L26"] = "=SUM(K16:K20)"
    ws["H34"] = "Resultado del Ejercicio"
    ws["K34"] = "=ER!H70"
    ws["H36"] = "TOTAL CAPITAL CONTABLE"
    ws["L36"] = "=SUM(K31:K34)"
    ws["H45"] = "TOTAL PASIVO+CAPITAL CONTABLE"
    ws["L45"] = "=L26+L36"
    ws["H47"] = "CUADRE"
    ws["L47"] = "=F45-L45"
    formula_cells.extend(
        (
            "BG!F26", "BG!F36", "BG!F42", "BG!F45", "BG!L26",
            "BG!K34", "BG!L36", "BG!L45", "BG!L47",
        )
    )

    _apply_versioned_sheet_presentation(
        ws, spec, white_bounds=(1, 47, 1, 12)
    )
    _configure_print_layout(ws, print_area="B7:L47", title_rows="7:10")


def _write_bal_sheet(
    ws,
    company: str,
    period: str,
    dataset: Any,
    metadata: Any,
    formula_cells: list[str],
) -> None:
    spec = _load_style_spec(BAL_STYLE_SPEC_PATH)
    rfc = _read_field(metadata, ("rfc",)) or "RFC por confirmar"
    ws["C1"] = company
    ws["C2"] = "Balanza de Comprobaci\u00f3n"
    ws["C3"] = _bal_period_caption(period)
    ws["C4"] = f"RFC: {rfc}"
    headers = ("CUENTA", "SALDO INICIAL", "DEBE", "HABER", "SALDO FINAL")
    for column, value in enumerate(headers, start=3):
        ws.cell(6, column, value)

    rows = _coerce_list_of_dicts(_read_field(dataset, ("rows", "all_rows")))
    accumulator_rows: list[int] = []
    for index, row in enumerate(rows, start=7):
        ws.cell(index, 3, row.get("account_raw") or row.get("account_code"))
        ws.cell(index, 4, row.get("saldo_inicial", 0))
        ws.cell(index, 5, row.get("debe", 0))
        ws.cell(index, 6, row.get("haber", 0))
        ws.cell(index, 7, row.get("saldo_final", 0))
        if row.get("is_accumulator"):
            accumulator_rows.append(index)
    separator_row = 7 + len(rows)
    sum_row = separator_row + 1
    ws.cell(sum_row, 3, "SUMAS IGUALES")
    ws.cell(sum_row, 4, "=0")
    formula_cells.append(f"BAL!D{sum_row}")
    for column in (5, 6):
        formula = _sum_accumulator_formula(get_column_letter(column), accumulator_rows)
        ws.cell(sum_row, column, formula)
        formula_cells.append(f"BAL!{get_column_letter(column)}{sum_row}")
    ws.cell(sum_row, 7, f"=D{sum_row}+E{sum_row}-F{sum_row}")
    formula_cells.append(f"BAL!G{sum_row}")

    _apply_bal_presentation(ws, rows, separator_row, sum_row, spec)
    ws.freeze_panes = None
    _configure_print_layout(
        ws, print_area=f"C1:G{sum_row}", title_rows="1:6"
    )


def _sum_accumulator_formula(column: str, rows: list[int]) -> str:
    if not rows:
        return "=0"
    return "=SUM(" + ",".join(f"{column}{row}" for row in rows) + ")"


def _apply_bal_presentation(
    ws,
    rows: list[dict[str, Any]],
    separator_row: int,
    sum_row: int,
    spec: Mapping[str, Any],
) -> None:
    _apply_sheet_geometry(ws, spec["geometry"])
    for coordinate, style_id in spec["cells"].items():
        row, _ = coordinate_to_tuple(coordinate)
        if row <= 6:
            _apply_style(ws[coordinate], spec["styles"][style_id])

    profiles = spec["dynamic_profiles"]
    for index, row_data in enumerate(rows, start=7):
        band = "light" if (index - 7) % 2 == 0 else "dark"
        weight = "bold" if row_data.get("is_accumulator") else "regular"
        profile = profiles[f"{band}_{weight}"]
        _apply_style(ws.cell(index, 3), spec["styles"][profile["label"]])
        for column in range(4, 8):
            _apply_style(ws.cell(index, column), spec["styles"][profile["numeric"]])

    for target_row, profile_name in (
        (separator_row, "separator"),
        (sum_row, "total"),
    ):
        profile = profiles[profile_name]
        _apply_style(ws.cell(target_row, 3), spec["styles"][profile["label"]])
        for column in range(4, 8):
            _apply_style(ws.cell(target_row, column), spec["styles"][profile["numeric"]])


def _apply_versioned_sheet_presentation(
    ws,
    spec: Mapping[str, Any],
    *,
    white_bounds: tuple[int, int, int, int] | None = None,
) -> None:
    _apply_sheet_geometry(ws, spec["geometry"])
    for coordinate, style_id in spec["cells"].items():
        _apply_style(ws[coordinate], spec["styles"][style_id])
    if white_bounds:
        min_row, max_row, min_col, max_col = white_bounds
        _apply_white_area(
            ws,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
        ws.sheet_view.showGridLines = False


def _configure_print_layout(ws, *, print_area: str, title_rows: str) -> None:
    ws.print_area = print_area
    ws.print_title_rows = title_rows
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = None


def _apply_white_area(ws, *, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    white = PatternFill(fill_type="solid", fgColor="FFFFFFFF", bgColor="FFFFFFFF")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = white


def _apply_sheet_geometry(ws, geometry: Mapping[str, Any]) -> None:
    ws.sheet_format.defaultRowHeight = geometry.get("default_row_height")
    ws.sheet_format.defaultColWidth = geometry.get("default_column_width")
    ws.sheet_format.baseColWidth = geometry.get("base_column_width")
    for column, width in geometry.get("column_widths", {}).items():
        ws.column_dimensions[column].width = width
        ws.column_dimensions[column].hidden = column in geometry.get("hidden_columns", [])
    for row, height in geometry.get("row_heights", {}).items():
        ws.row_dimensions[int(row)].height = height
    for row in geometry.get("hidden_rows", []):
        ws.row_dimensions[int(row)].hidden = True
    existing_merges = {str(item) for item in ws.merged_cells.ranges}
    for merged_range in geometry.get("merged_ranges", []):
        if merged_range not in existing_merges:
            ws.merge_cells(merged_range)
    for name, value in geometry.get("page_margins", {}).items():
        setattr(ws.page_margins, name, value)
    for name, value in geometry.get("page_setup", {}).items():
        if value is not None:
            setattr(ws.page_setup, name, value)
    for name, value in geometry.get("page_setup_properties", {}).items():
        if value is not None:
            setattr(ws.sheet_properties.pageSetUpPr, name, value)
    for name, value in geometry.get("print_options", {}).items():
        if value is not None:
            setattr(ws.print_options, name, value)
    for name, value in geometry.get("sheet_view", {}).items():
        if value is not None:
            setattr(ws.sheet_view, name, value)
    ws.freeze_panes = geometry.get("freeze_panes")


def _load_style_spec(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def _configure_recalculation_on_open(workbook: Workbook) -> None:
    """Ask Excel-compatible clients to fully recalculate formula cells."""

    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    for attribute, value in (
        ("calcMode", "auto"),
        ("fullCalcOnLoad", True),
        ("forceFullCalc", True),
    ):
        if hasattr(calculation, attribute):
            setattr(calculation, attribute, value)


def save_er_workbook(
    dataset: Any | None = None,
    output_path: str | Path | None = None,
    *,
    metadata: Any | None = None,
    source_path: str | Path | None = None,
) -> WorkbookBuildResult:
    """Compatibility wrapper for :func:`save_financial_statements_workbook`."""

    return save_financial_statements_workbook(
        dataset, output_path, metadata=metadata, source_path=source_path
    )


def save_financial_statements_workbook(
    er_dataset: Any | None = None,
    output_path: str | Path | None = None,
    *,
    bg_dataset: Any | None = None,
    bal_dataset: Any | None = None,
    metadata: Any | None = None,
    source_path: str | Path | None = None,
) -> WorkbookBuildResult:
    """Build and save a self-contained financial-statements workbook."""

    if bg_dataset is None or bal_dataset is None:
        derived_bg, derived_bal, metadata = _derive_supporting_datasets(
            er_dataset, metadata=metadata, source_path=source_path
        )
        bg_dataset = bg_dataset or derived_bg
        bal_dataset = bal_dataset or derived_bal
    result = build_financial_statements_workbook(
        er_dataset,
        bg_dataset=bg_dataset,
        bal_dataset=bal_dataset,
        metadata=metadata,
        source_path=source_path,
    )
    target = Path(output_path) if output_path is not None else _default_output_path(result.company, result.period)
    target.parent.mkdir(parents=True, exist_ok=True)
    result.workbook.save(target)
    result.output_path = target
    return result


def detect_source_metadata(path: str | Path) -> dict[str, str | None]:
    """Read lightweight company/period metadata from a workbook-like source.

    This is intentionally not a full balanza parser; it only supports local
    evidence generation until the parser agent wires `parse_balanza`.
    """

    source = Path(path)
    with source.open("rb") as handle:
        wb = load_workbook(handle, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        company = None
        period = None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
            for value in row:
                if value is None:
                    continue
                text = str(value).strip()
                if not text:
                    continue
                lowered = text.lower()
                if company is None and "periodo" not in lowered and re.search(r"\b[A-Z0-9]{12,13}\b", text):
                    company = text
                if period is None:
                    match = re.search(r"(\d{4})[-/](\d{1,2})", text)
                    if match:
                        period = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
            if company and period:
                break
        wb.close()
    return {"company": company, "period": period, "source_path": str(source)}


def _write_header(ws, company: str, period: str) -> None:
    ws["B9"] = company
    ws["B10"] = "Estado de Resultados"
    ws["B11"] = _period_caption(period)
    ws["B12"] = "(Importes expresados en pesos)"
    ws["D15"] = "Del Período "
    ws["F15"] = "%"
    ws["H15"] = "Acumulado"
    ws["J15"] = "%"


def _write_er_body(ws, line_values: dict[str, dict[str, Any]], warnings: list[str]) -> list[str]:
    formula_cells: list[str] = []
    numeric_rows: list[int] = []

    for spec in ER_LAYOUT:
        row = int(spec["row"])
        label = str(spec["label"])
        kind = spec.get("kind", "line")
        ws.cell(row=row, column=2, value=label)

        if kind == "section":
            continue

        numeric_rows.append(row)
        if kind == "subtotal":
            formula = "=" + str(spec["formula"]).format(col="H")
            _set_formula(ws, f"H{row}", formula, formula_cells)
            continue

        values = _line_amounts(line_values.get(str(spec.get("key"))), warnings, label)
        ws[f"H{row}"] = values["accumulated"]

    for row in numeric_rows:
        _set_formula(ws, f"J{row}", f'=IF($H$18=0,0,H{row}/$H$18)', formula_cells)

    return formula_cells


def _set_formula(ws, cell: str, formula: str, formula_cells: list[str]) -> None:
    safe_formula = _sanitize_formula(formula)
    ws[cell] = safe_formula
    formula_cells.append(f"ER!{cell}")


def _sanitize_formula(formula: str) -> str:
    clean = formula.strip()
    if not clean.startswith("="):
        clean = "=" + clean
    upper = clean.upper()
    if any(token in upper for token in ERROR_TOKENS) or "[" in clean or "]" in clean:
        return "=0"
    return clean


def _style_er_sheet(ws) -> None:
    """Apply the exact non-fill ER contract and the required white canvas."""

    spec = _load_er_style_spec()
    _apply_versioned_sheet_presentation(
        ws, spec, white_bounds=(1, 70, 1, 10)
    )
    _configure_print_layout(ws, print_area="B9:J70", title_rows="9:15")


def _load_er_style_spec() -> dict[str, Any]:
    with ER_STYLE_SPEC_PATH.open(encoding="utf-8") as handle:
        return json.load(handle)


def _color_from_spec(value: dict[str, Any] | None) -> Color | None:
    if not value:
        return None
    kwargs: dict[str, Any] = {"type": value["type"]}
    for key in ("rgb", "indexed", "theme", "tint"):
        if value.get(key) is not None:
            kwargs[key] = value[key]
    return Color(**kwargs)


def _side_from_spec(value: dict[str, Any] | None) -> Side | None:
    if not value:
        return None
    return Side(style=value.get("style"), color=_color_from_spec(value.get("color")))


def _apply_style(cell: Any, spec: dict[str, Any]) -> None:
    font = spec["font"]
    cell.font = Font(
        name=font["name"],
        sz=font["size"],
        b=font["bold"],
        i=font["italic"],
        underline=font["underline"],
        strike=font["strike"],
        color=_color_from_spec(font["color"]),
        vertAlign=font["vertAlign"],
        charset=font["charset"],
        family=font["family"],
        scheme=font["scheme"],
        outline=font["outline"],
        shadow=font["shadow"],
        condense=font["condense"],
        extend=font["extend"],
    )
    alignment = spec["alignment"]
    cell.alignment = Alignment(
        horizontal=alignment["horizontal"],
        vertical=alignment["vertical"],
        textRotation=alignment["textRotation"],
        wrap_text=alignment["wrapText"],
        shrink_to_fit=alignment["shrinkToFit"],
        indent=alignment["indent"],
        relativeIndent=alignment["relativeIndent"],
        justifyLastLine=alignment["justifyLastLine"],
        readingOrder=alignment["readingOrder"],
    )
    fill = spec["fill"]
    cell.fill = PatternFill(
        fill_type=fill["fillType"],
        fgColor=_color_from_spec(fill["fgColor"]),
        bgColor=_color_from_spec(fill["bgColor"]),
    )
    border = spec["border"]
    cell.border = Border(
        left=_side_from_spec(border["left"]),
        right=_side_from_spec(border["right"]),
        top=_side_from_spec(border["top"]),
        bottom=_side_from_spec(border["bottom"]),
        start=_side_from_spec(border.get("start")),
        end=_side_from_spec(border.get("end")),
        diagonal=_side_from_spec(border["diagonal"]),
        diagonalUp=border["diagonalUp"],
        diagonalDown=border["diagonalDown"],
        outline=border["outline"],
        vertical=_side_from_spec(border["vertical"]),
        horizontal=_side_from_spec(border["horizontal"]),
    )
    protection = spec["protection"]
    cell.protection = Protection(
        locked=protection["locked"], hidden=protection["hidden"]
    )
    cell.number_format = spec["numberFormat"]


def _metadata_from_inputs(dataset: Any, metadata: Any, source_path: str | Path | None) -> dict[str, Any]:
    warnings: list[str] = []
    values: dict[str, Any] = {}
    for source in (metadata, dataset):
        if source is None:
            continue
        for key in ("company", "empresa", "company_name", "nombre_empresa"):
            _set_if_present(values, "company", _read_field(source, (key,)))
        for key in ("period", "periodo", "period_detected", "periodo_detectado"):
            _set_if_present(values, "period", _normalize_period(_read_field(source, (key,))))
        for key in ("source_path", "archivo_origen", "input_path"):
            _set_if_present(values, "source_path", _read_field(source, (key,)))

    if source_path is not None:
        values["source_path"] = str(source_path)
        try:
            detected = detect_source_metadata(source_path)
        except Exception as exc:  # pragma: no cover - defensive integration path
            warnings.append(f"Could not detect source metadata from {source_path}: {exc}")
        else:
            _set_if_present(values, "company", detected.get("company"))
            _set_if_present(values, "period", detected.get("period"))

    company = str(values.get("company") or "Empresa por confirmar").strip()
    period = str(values.get("period") or "Periodo por confirmar").strip()
    if company == "Empresa por confirmar":
        warnings.append("Company was not provided by parser/dataset metadata.")
    if period == "Periodo por confirmar":
        warnings.append("Period was not provided by parser/dataset metadata.")
    return {"company": company, "period": period, "source_path": values.get("source_path"), "warnings": warnings}


def _index_dataset_lines(dataset: Any, warnings: list[str]) -> dict[str, dict[str, Any]]:
    lines = _read_field(dataset, ("lines", "lineas", "items", "rows", "renglones"))
    indexed: dict[str, dict[str, Any]] = {}
    if lines is None:
        return indexed
    if isinstance(lines, dict):
        iterable: Iterable[Any] = lines.items()
    else:
        iterable = lines if isinstance(lines, Iterable) and not isinstance(lines, (str, bytes)) else []
    for raw_line in iterable:
        if isinstance(raw_line, tuple) and len(raw_line) == 2:
            line_key, raw_line = raw_line
        else:
            line_key = None
        line = _to_dict(raw_line)
        if not line and line_key is not None:
            line = {"key": line_key, "amount": raw_line}
        elif line_key is not None:
            line.setdefault("key", line_key)
        label = _first_present(line, ("key", "id", "label", "concept", "concepto", "name", "nombre"))
        if label is None:
            continue
        indexed[_slug(str(label))] = line
    return indexed


def _line_amounts(line: dict[str, Any] | None, warnings: list[str], label: str) -> dict[str, float]:
    if line is None:
        return {"period": 0.0, "accumulated": 0.0}
    period_formula = _first_present(line, ("period_formula", "formula_periodo", "formula"))
    accumulated_formula = _first_present(line, ("accumulated_formula", "formula_acumulado"))
    if period_formula or accumulated_formula:
        warnings.append(f"Dataset formulas for '{label}' were ignored; generated ER uses safe internal formulas only.")
    period = _number(_first_present(line, ("period_amount", "amount", "importe_periodo", "periodo", "value", "valor")))
    accumulated = _number(_first_present(line, ("accumulated_amount", "accumulated", "acumulado", "importe_acumulado")))
    if accumulated is None:
        accumulated = period
    return {"period": period or 0.0, "accumulated": accumulated or 0.0}


def _read_field(source: Any, keys: tuple[str, ...]) -> Any:
    if source is None:
        return None
    mapping = _to_dict(source)
    for key in keys:
        if key in mapping and mapping[key] not in (None, ""):
            return mapping[key]
    for key in keys:
        if hasattr(source, key):
            value = getattr(source, key)
            if value not in (None, ""):
                return value
    return None


def _to_dict(value: Any) -> dict[str, Any]:
    if value is None:
        return {}
    if isinstance(value, dict):
        return value
    if is_dataclass(value):
        return {field_name: getattr(value, field_name) for field_name in value.__dataclass_fields__}
    if hasattr(value, "__dict__"):
        return vars(value)
    return {}


def _first_present(mapping: dict[str, Any], keys: tuple[str, ...]) -> Any:
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return None


def _set_if_present(values: dict[str, Any], key: str, value: Any) -> None:
    if values.get(key) in (None, "") and value not in (None, ""):
        values[key] = value


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None


def _normalize_period(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    match = re.search(r"(\d{4})[-/](\d{1,2})", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    return text


def _period_caption(period: str) -> str:
    match = re.search(r"(\d{4})-(\d{2})", period)
    if not match:
        return f"Periodo: {period}"
    year = int(match.group(1))
    month = int(match.group(2))
    last_day = calendar.monthrange(year, month)[1]
    month_name = MONTHS_ES.get(month, str(month))
    return f"Del 1ro de Enero al {last_day} de {month_name} {year}"


def _bg_period_caption(period: str) -> str:
    match = re.search(r"(\d{4})-(\d{2})", period)
    if not match:
        return f"Al cierre de {period}"
    year = int(match.group(1))
    month = int(match.group(2))
    return f"Al {calendar.monthrange(year, month)[1]} de {MONTHS_ES.get(month, month)} de {year}"


def _bal_period_caption(period: str) -> str:
    match = re.search(r"(\d{4})-(\d{2})", period)
    if not match:
        return period.upper()
    return f"{MONTHS_ES.get(int(match.group(2)), match.group(2)).upper()}.{match.group(1)}"


def _slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "_", ascii_text.lower()).strip("_")


def _coerce_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [value]
    if isinstance(value, Iterable):
        return [str(item) for item in value]
    return [str(value)]


def _coerce_list_of_dicts(value: Any) -> list[dict[str, Any]]:
    if value is None:
        return []
    if isinstance(value, Mapping):
        return [dict(value)]
    if not isinstance(value, Iterable) or isinstance(value, (str, bytes)):
        return []
    return [_to_dict(item) for item in value]


def _default_output_path(company: str, period: str) -> Path:
    company_slug = _slug(company)[:60] or "empresa"
    period_slug = _slug(period) or "periodo"
    return DEFAULT_OUTPUT_DIR / f"estados_financieros_{company_slug}_{period_slug}.xlsx"
